#!/usr/bin/env python3
import csv, json, os, re, sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook

run_dir, template_path, source_path, output_path = map(Path, sys.argv[1:5])

XHS_ACCOUNTS = [
    {"profile":"account-02","name":"广州大学城租房-研寓","operator":"孝西","team":"管家团队"},
    {"profile":"account-03","name":"广州研寓租房大学城","operator":"嘉明","team":"管家团队"},
    {"profile":"account-04","name":"大学城捞房长短租随意","operator":"梦琪","team":"运营团队"},
    {"profile":"account-05","name":"暴走大学城探房版","operator":"淼淼","team":"运营团队"},
    {"profile":"account-06","name":"广州大学城-研舍公寓","operator":"紫莹","team":"管家团队"},
    {"profile":"account-07","name":"广州大学城租房-维特","operator":"珂珂","team":"管家团队"},
    {"profile":"account-08","name":"大学城租房 | 研舍","operator":"传坤","team":"管家团队"},
    {"profile":"account-09","name":"番禺大学城租房-尚维特","operator":"余路","team":"管家团队"},
]

def xhs_owner_team(name):
    text=str(name or "")
    return "运营团队" if "大学城捞房" in text or "暴走大学城" in text else "管家团队"

def weekly_lead_targets(opened, copied):
    return {
        week_key: {"opened": opened[index], "copied": copied[index]}
        for index, week_key in enumerate(("w1","w2","w3","w4","we"))
    }

XHS_LEAD_TARGET_NAMES = {
    "account-02": "广州大学城租房-研寓",
    "account-03": "广州研寓租房大学城",
    "account-04": "大学城捞房",
    "account-05": "暴走大学城",
    "account-06": "广州大学城—研舍公寓",
    "account-07": "大学城租房-小梦",
    "account-08": "大学城租房｜研舍",
    "account-09": "尚维特",
}

XHS_LEAD_TARGETS = {
    "2026-08": {
        "account-02": weekly_lead_targets([16,24,14,16,5], [10,14,8,9,3]),
        "account-03": weekly_lead_targets([74,71,92,87,14], [55,52,68,64,10]),
        "account-04": weekly_lead_targets([68,91,135,136,42], [50,67,99,99,31]),
        "account-05": weekly_lead_targets([62,90,54,103,9], [48,70,42,80,7]),
        "account-06": weekly_lead_targets([16,26,21,29,8], [10,17,14,19,5]),
        "account-07": weekly_lead_targets([17,10,20,17,6], [12,7,14,12,4]),
        "account-08": weekly_lead_targets([13,21,69,81,14], [0,0,1,1,0]),
        "account-09": weekly_lead_targets([48,40,31,32,13], [34,29,22,23,9]),
    },
}

def attach_xhs_lead_targets(payload):
    month=payload.get("month","")
    configured=XHS_LEAD_TARGETS.get(month,{})
    accounts=[]
    for account in payload.get("accounts",[]):
        profile=account.get("profile","")
        accounts.append({
            **account,
            "targetName":XHS_LEAD_TARGET_NAMES.get(profile,account.get("name",profile)),
            "targets":configured.get(profile,{}),
        })
    return {
        **payload,
        "targetMonth":month if configured else "",
        "targetWeeks":["w1","w2","w3","w4","we"] if configured else [],
        "accounts":accounts,
    }

def latest_xhs_summary():
    """Locate the newest collector summary without coupling the dashboard to a dated folder."""
    candidates=[]
    configured=os.environ.get("XHS_CONTENT_WEEKLY_JSON","").strip()
    if configured: candidates.append(Path(configured))
    roots=[Path("/home/ubuntu/xhs-account-isolation/data"),Path("/opt/xhs-account-isolation/data")]
    for root in roots:
        if root.exists(): candidates.extend(root.glob("content-stats*/content-weekly-summary-latest.json"))
    existing=[path for path in candidates if path.is_file()]
    return max(existing,key=lambda path:path.stat().st_mtime) if existing else None

def build_xhs_content(fallback):
    summary_path=latest_xhs_summary()
    if not summary_path: return fallback or {"generatedAt":"","month":"","weeks":[],"accounts":XHS_ACCOUNTS,"dailyReading":[]}
    summary=json.loads(summary_path.read_text(encoding="utf-8"))
    end_date=datetime.strptime(summary["period"]["end_date"],"%Y-%m-%d").date()
    month_prefix=str(end_date.month)
    week_rows=[row for row in summary.get("weekly_totals",[]) if row.get("week_label","").startswith(month_prefix+"W") and row.get("week_start","")<=end_date.isoformat()]
    weeks=[{"key":row["week_label"][len(month_prefix):].lower(),"label":row["week_label"][len(month_prefix):],"start":row["week_start"],"end":row["week_end"]} for row in week_rows]
    by_profile_week={(row.get("profile"),row.get("week_label")):row for row in summary.get("account_week_rows",[])}
    accounts=[]
    for account in XHS_ACCOUNTS:
        metrics={}
        for week in weeks:
            source=by_profile_week.get((account["profile"],month_prefix+week["label"]))
            metrics[week["key"]]={
                "notes":source.get("note_count") if source else None,
                "reads":source.get("note_reading_count") if source else None,
                "interactions":source.get("interaction_count") if source else None,
                "interactionRate":source.get("interaction_rate") if source else None,
            }
        accounts.append({**account,"metrics":metrics})
    daily_totals=defaultdict(int)
    account_daily=[]
    daily_path=summary_path.with_name("daily-reading-latest.csv")
    if daily_path.is_file():
        with daily_path.open(encoding="utf-8-sig",newline="") as handle:
            for row in csv.DictReader(handle):
                if row.get("status") == "ok" and row.get("date"):
                    reading_count=int(row.get("reading_count") or 0)
                    daily_totals[row["date"]]+=reading_count
                    account=next((item for item in XHS_ACCOUNTS if item["profile"]==row.get("profile")),None)
                    account_daily.append({"profile":row.get("profile", ""),"accountName":(account or {}).get("name") or row.get("account_name") or row.get("profile", ""),"date":row["date"],"readingCount":reading_count})
    account_daily.sort(key=lambda row:(row["profile"],row["date"]))
    return {"generatedAt":summary.get("generated_at","")[:19].replace("T"," "),"month":f"{end_date.year:04d}-{end_date.month:02d}","weeks":weeks,"accounts":accounts,"dailyReading":[{"date":date,"readingCount":count} for date,count in sorted(daily_totals.items(),reverse=True)],"accountDailyReading":account_daily}

def latest_xhs_note_published_csv():
    """Locate the immutable per-account note publication history."""
    candidates=[]
    configured=os.environ.get("XHS_NOTE_PUBLISHED_CSV","").strip()
    if configured: candidates.append(Path(configured))
    roots=[Path("/home/ubuntu/xhs-account-isolation/data"),Path("/opt/xhs-account-isolation/data")]
    for root in roots:
        candidate=root / "immutable-history" / "note-published-daily.csv"
        if candidate.is_file(): candidates.append(candidate)
    existing=[path for path in candidates if path.is_file()]
    return max(existing,key=lambda path:path.stat().st_mtime) if existing else None

def build_xhs_note_published(fallback):
    history_path=latest_xhs_note_published_csv()
    if not history_path:
        return fallback or {"generatedAt":"","rows":[],"accounts":[]}
    configured_accounts={account["profile"]:account for account in XHS_ACCOUNTS}
    type_counts=defaultdict(lambda:{"图文":0,"视频":0,"待识别":0})
    notes_path=history_path.parent / "notes.csv"
    if notes_path.is_file():
        with notes_path.open(encoding="utf-8-sig",newline="") as handle:
            for note in csv.DictReader(handle):
                profile=str(note.get("profile") or "").strip()
                published_date=str(note.get("published_date") or "").strip()
                if not profile or not re.fullmatch(r"\d{4}-\d{2}-\d{2}",published_date):
                    continue
                note_type=str(note.get("note_type") or "").strip()
                if note_type not in {"图文","视频"}: note_type="待识别"
                type_counts[(profile,published_date)][note_type]+=1
    rows=[]
    with history_path.open(encoding="utf-8-sig",newline="") as handle:
        for source in csv.DictReader(handle):
            profile=str(source.get("profile") or "").strip()
            published_date=str(source.get("date") or "").strip()
            if not profile or not re.fullmatch(r"\d{4}-\d{2}-\d{2}",published_date):
                continue
            configured=configured_accounts.get(profile,{})
            account_name=str(source.get("account_name") or configured.get("name") or profile).strip()
            published_count=int(source.get("note_count") or 0)
            counts=type_counts[(profile,published_date)]
            graphic_count=int(counts["图文"])
            video_count=int(counts["视频"])
            classified_count=graphic_count+video_count
            if classified_count>published_count:
                raise RuntimeError(f"XHS note type counts exceed published total: {profile} {published_date}")
            pending_count=published_count-classified_count
            rows.append({"profile":profile,"accountName":account_name,"publishedDate":published_date,"graphicCount":graphic_count,"videoCount":video_count,"pendingCount":pending_count,"publishedCount":published_count})
    rows.sort(key=lambda row:(row["publishedDate"],row["accountName"]),reverse=True)
    accounts=[]
    for profile in sorted({row["profile"] for row in rows}):
        configured=configured_accounts.get(profile,{})
        account_name=next((row["accountName"] for row in rows if row["profile"]==profile),configured.get("name") or profile)
        accounts.append({"profile":profile,"name":account_name})
    accounts.sort(key=lambda row:row["name"])
    return {"generatedAt":datetime.fromtimestamp(history_path.stat().st_mtime).astimezone().strftime("%Y-%m-%d %H:%M"),"rows":rows,"accounts":accounts}

def latest_xhs_lead_summary():
    """Locate the newest weekly lead summary produced by the lead collector."""
    candidates=[]
    configured=os.environ.get("XHS_LEAD_WEEKLY_JSON","").strip()
    if configured: candidates.append(Path(configured))
    roots=[Path("/home/ubuntu/xhs-account-isolation/data"),Path("/opt/xhs-account-isolation/data")]
    for root in roots:
        if root.exists(): candidates.extend(root.glob("lead-stats*/weekly-summary-latest.json"))
    existing=[path for path in candidates if path.is_file()]
    return max(existing,key=lambda path:path.stat().st_mtime) if existing else None

def build_customer_data(fallback):
    """Build the operations-team seven-day cross-source customer-service funnel."""
    empty={"generatedAt":"","startDate":"","endDate":"","accounts":[],"dailyRows":[],"totals":{},"funnel":{},"sourceNote":""}
    roots=[Path("/home/ubuntu/xhs-account-isolation/data"),Path("/opt/xhs-account-isolation/data")]
    configured_root=os.environ.get("XHS_DATA_ROOT","").strip()
    if configured_root: roots.insert(0,Path(configured_root))
    history_root=next((root/"immutable-history" for root in roots if (root/"immutable-history").is_dir()),None)
    customer_path=Path(os.environ.get("CUSTOMER_SERVICE_DATA_XLSX","/opt/yuxiaor-automation/manual-data/operations-customer-data-2026-08.xlsx"))
    required_files=("note-published-daily.csv","professional-reading-daily.csv","lead-daily.csv")
    if not history_root or not customer_path.is_file() or any(not (history_root/name).is_file() for name in required_files):
        return fallback or empty
    operations_accounts=[account for account in XHS_ACCOUNTS if account["team"]=="运营团队"]
    profiles={account["profile"] for account in operations_accounts}

    def read_daily_csv(name,fields):
        values=defaultdict(lambda:defaultdict(int)); coverage=defaultdict(set)
        with (history_root/name).open(encoding="utf-8-sig",newline="") as handle:
            for row in csv.DictReader(handle):
                profile=str(row.get("profile") or "").strip(); date_text=str(row.get("date") or "").strip()
                if profile not in profiles or not re.fullmatch(r"\d{4}-\d{2}-\d{2}",date_text): continue
                coverage[date_text].add(profile)
                for field in fields: values[date_text][field]+=int(float(row.get(field) or 0))
        return values,{date for date,seen in coverage.items() if seen==profiles}

    published,_=read_daily_csv("note-published-daily.csv",("note_count",))
    reading,reading_dates=read_daily_csv("professional-reading-daily.csv",("reading_count",))
    leads,lead_dates=read_daily_csv("lead-daily.csv",("inbound_users","personal_wechat_copies"))
    workbook=load_workbook(customer_path,read_only=True,data_only=True)
    sheet=workbook["8月数据"] if "8月数据" in workbook.sheetnames else workbook.active
    header={str(cell.value or "").strip():cell.column for cell in sheet[4]}
    def header_column(prefix): return next((column for name,column in header.items() if name.startswith(prefix)),None)
    columns={"date":header_column("日期"),"wechatAdds":header_column("微信添加（小红书"),"actualTours":header_column("看房（实际"),"signed":header_column("成交"),"deposits":header_column("定金")}
    if any(column is None for column in columns.values()): return fallback or empty
    customer={}
    for row in sheet.iter_rows(min_row=5,values_only=True):
        if len(row)<max(columns.values()): continue
        raw_date=row[columns["date"]-1]
        if isinstance(raw_date,datetime): date_text=raw_date.date().isoformat()
        else:
            try: date_text=datetime.strptime(str(raw_date or "").strip(),"%Y/%m/%d").date().isoformat()
            except ValueError: continue
        customer[date_text]={key:int(float(row[column-1] or 0)) for key,column in columns.items() if key!="date"}
    dates=sorted(set(customer)&reading_dates&lead_dates)[-7:]
    daily=[{"date":date_text,"published":published[date_text]["note_count"],"reading":reading[date_text]["reading_count"],"inbound":leads[date_text]["inbound_users"],"leads":leads[date_text]["personal_wechat_copies"],**customer[date_text]} for date_text in dates]
    fields=("published","reading","inbound","leads","wechatAdds","actualTours","signed","deposits")
    totals={field:sum(row[field] for row in daily) for field in fields}
    def ratio(numerator,denominator): return numerator/denominator if denominator else None
    funnel={"readsPerPost":ratio(totals.get("reading",0),totals.get("published",0)),"readingToInbound":ratio(totals.get("inbound",0),totals.get("reading",0)),"inboundToLeads":ratio(totals.get("leads",0),totals.get("inbound",0)),"leadsToWechat":ratio(totals.get("wechatAdds",0),totals.get("leads",0)),"wechatToActualTours":ratio(totals.get("actualTours",0),totals.get("wechatAdds",0)),"actualToursToDeposits":ratio(totals.get("deposits",0),totals.get("actualTours",0)),"depositsToSigned":ratio(totals.get("signed",0),totals.get("deposits",0)),"actualToursToSigned":ratio(totals.get("signed",0),totals.get("actualTours",0)),"readingToSigned":ratio(totals.get("signed",0),totals.get("reading",0))}
    newest=max([customer_path.stat().st_mtime]+[(history_root/name).stat().st_mtime for name in required_files])
    return {"generatedAt":datetime.fromtimestamp(newest).astimezone().strftime("%Y-%m-%d %H:%M"),"startDate":dates[0] if dates else "","endDate":dates[-1] if dates else "","accounts":operations_accounts,"dailyRows":daily,"totals":totals,"funnel":funnel,"sourceNote":"运营团队专业号与客服表格共同覆盖的最近7个自然日；同日跨来源汇总，不代表按客户ID追踪。"}

def build_meter_management(fallback):
    """Attach the last complete, sanitized meter snapshot and latest run status."""
    data_path=Path(os.environ.get("METER_MANAGEMENT_JSON","/opt/yuxiaor-automation/data/meter-management/latest.json"))
    status_path=Path(os.environ.get("METER_MANAGEMENT_STATUS_JSON","/opt/yuxiaor-automation/data/meter-management/status.json"))
    empty={"schemaVersion":1,"source":"微亭易租设备管理","projectId":"","collectedAt":"","summary":{"total":0,"online":0,"offline":0,"negative":0,"keepElectric":0},"keepElectricDevices":[],"negativeDevices":[],"offlineDevices":[],"collectionStatus":{"state":"missing"}}
    content=dict(fallback or empty)
    if data_path.is_file():
        loaded=json.loads(data_path.read_text(encoding="utf-8"))
        if not isinstance(loaded,dict): raise RuntimeError("Meter-management snapshot must be an object")
        content=loaded
    status={"state":"missing"}
    if status_path.is_file():
        loaded_status=json.loads(status_path.read_text(encoding="utf-8"))
        if isinstance(loaded_status,dict): status=loaded_status
    content["collectionStatus"]=status
    summary=content.get("summary",{})
    keep_rows=content.get("keepElectricDevices",[])
    negative_rows=content.get("negativeDevices",[])
    offline_rows=content.get("offlineDevices",[])
    if not all(isinstance(rows,list) for rows in (keep_rows,negative_rows,offline_rows)):
        raise RuntimeError("Meter-management exception lists must be arrays")
    if int(summary.get("keepElectric") or 0)!=len(keep_rows) or int(summary.get("negative") or 0)!=len(negative_rows) or int(summary.get("offline") or 0)!=len(offline_rows):
        raise RuntimeError("Meter-management summary does not reconcile")
    allowed={"deviceId","deviceName","areaName","onlineStatus","powerStatus","remainingPower","updatedAt","keepElectric"}
    if any(set(row)-allowed for rows in (keep_rows,negative_rows,offline_rows) for row in rows):
        raise RuntimeError("Meter-management payload contains an unexpected device field")
    if any(not isinstance(row.get("keepElectric"),bool) for rows in (keep_rows,negative_rows,offline_rows) for row in rows):
        raise RuntimeError("Meter-management keep-electric state must be boolean")
    return content

def latest_xhs_ad_immutable_history():
    """Locate the append-only Aurora per-note daily history used by the dashboard."""
    candidates=[]
    configured=os.environ.get("XHS_AD_IMMUTABLE_CSV","").strip()
    if configured: candidates.append(Path(configured))
    for root in (Path("/home/ubuntu/xhs-account-isolation/data"),Path("/opt/xhs-account-isolation/data")):
        candidate=root / "immutable-history" / "ad-note-daily.csv"
        if candidate.is_file(): candidates.append(candidate)
    existing=[path for path in candidates if path.is_file()]
    return max(existing,key=lambda path:path.stat().st_mtime) if existing else None

def build_xhs_ad_flow_from_immutable(history_path):
    """Aggregate append-only note history without relying on a separately rebuilt snapshot."""
    configured={account["profile"]:account for account in XHS_ACCOUNTS}
    source_rows=[]
    collected_times=[]
    with history_path.open(encoding="utf-8-sig",newline="") as handle:
        reader=csv.DictReader(handle)
        required={"date","profile","note_id","spend","private_message_opens","private_message_leads","owner_account_name","owner_user_id","owner_status","first_collected_at"}
        if not reader.fieldnames or not required.issubset(reader.fieldnames):
            raise RuntimeError("XHS ad immutable history columns invalid")
        for row in reader:
            report_date=str(row.get("date") or "").strip()
            profile=str(row.get("profile") or "").strip()
            note_id=str(row.get("note_id") or "").strip()
            if not re.fullmatch(r"\d{4}-\d{2}-\d{2}",report_date) or profile not in configured or not note_id:
                continue
            source_rows.append(row)
            collected_at=str(row.get("first_collected_at") or "").strip()
            if collected_at: collected_times.append(collected_at)
    if not source_rows:
        raise RuntimeError("XHS ad immutable history is empty")

    dates=sorted({row["date"].strip() for row in source_rows})
    first_date=datetime.strptime(dates[0],"%Y-%m-%d").date()
    last_date=datetime.strptime(dates[-1],"%Y-%m-%d").date()
    all_dates=[]
    cursor=first_date
    while cursor<=last_date:
        all_dates.append(cursor.isoformat())
        cursor+=timedelta(days=1)

    account_groups=defaultdict(lambda:{"noteCount":0,"spend":0.0,"opened":0,"leads":0})
    owner_groups={}
    unresolved_count=0
    for row in source_rows:
        report_date=row["date"].strip()
        profile=row["profile"].strip()
        spend=float(row.get("spend") or 0)
        opened=int(row.get("private_message_opens") or 0)
        leads=int(row.get("private_message_leads") or 0)
        account=account_groups[(report_date,profile)]
        account["noteCount"]+=1
        account["spend"]+=spend
        account["opened"]+=opened
        account["leads"]+=leads

        owner_name=str(row.get("owner_account_name") or "").strip()
        owner_user_id=str(row.get("owner_user_id") or "").strip()
        confirmed=bool(owner_name and str(row.get("owner_status") or "").strip()=="confirmed")
        if not confirmed: unresolved_count+=1
        owner_key=(report_date,owner_user_id or owner_name or "unresolved")
        owner=owner_groups.setdefault(owner_key,{
            "date":report_date,"ownerAccountName":owner_name,"ownerUserId":owner_user_id,
            "team":xhs_owner_team(owner_name),"noteCount":0,"spend":0.0,"opened":0,"leads":0,
            "ownerStatus":"confirmed" if confirmed else "unresolved",
            "ownerStatusLabel":"已确认" if confirmed else "待确认",
        })
        owner["noteCount"]+=1
        owner["spend"]+=spend
        owner["opened"]+=opened
        owner["leads"]+=leads
        if not confirmed:
            owner["ownerStatus"]="unresolved"
            owner["ownerStatusLabel"]="待确认"

    account_rows=[]
    for report_date in all_dates:
        for profile,configured_account in configured.items():
            values=account_groups[(report_date,profile)]
            spend=round(values["spend"],2)
            opened=values["opened"]
            leads=values["leads"]
            account_rows.append({
                "date":report_date,"profile":profile,"accountName":configured_account["name"],"team":configured_account["team"],
                "noteCount":values["noteCount"],"spend":spend,"opened":opened,
                "averageOpenCost":round(spend/opened,2) if opened else None,
                "leads":leads,"averageLeadCost":round(spend/leads,2) if leads else None,
                "status":"ok","statusLabel":"有数据" if values["noteCount"] else "已确认零消耗","error":"",
            })
    account_rows.sort(key=lambda row:(row["date"],row["profile"]),reverse=True)

    owner_rows=list(owner_groups.values())
    for row in owner_rows:
        row["spend"]=round(row["spend"],2)
        row["averageOpenCost"]=round(row["spend"]/row["opened"],2) if row["opened"] else None
        row["averageLeadCost"]=round(row["spend"]/row["leads"],2) if row["leads"] else None
    owner_rows.sort(key=lambda row:(row["date"],row["spend"],row["ownerAccountName"]),reverse=True)
    generated_at=max(collected_times)[:19].replace("T"," ") if collected_times else datetime.fromtimestamp(history_path.stat().st_mtime).astimezone().strftime("%Y-%m-%d %H:%M")
    return {
        "generatedAt":generated_at,"date":all_dates[-1],"periodLabel":f"{all_dates[0]} 至 {all_dates[-1]}","aggregation":"DAY",
        "historySource":"immutable-history/ad-note-daily.csv","historyMaxDate":all_dates[-1],"historyRowCount":len(source_rows),
        "accountRows":account_rows,"ownerRows":owner_rows,
        "ownerAccountCount":len({row["ownerUserId"] or row["ownerAccountName"] for row in owner_rows if row["ownerUserId"] or row["ownerAccountName"]}),
        "totalNoteCount":len(source_rows),"totalSpend":round(sum(float(row.get("spend") or 0) for row in source_rows),2),
        "totalOpened":sum(int(row.get("private_message_opens") or 0) for row in source_rows),
        "totalLeads":sum(int(row.get("private_message_leads") or 0) for row in source_rows),
        "unresolvedOwnerCount":unresolved_count,
    }

def latest_xhs_ad_note_summary():
    """Locate the newest official Aurora note-ad snapshot synchronized to the workbench."""
    candidates=[]
    configured=os.environ.get("XHS_AD_NOTE_JSON","").strip()
    if configured: candidates.append(Path(configured))
    roots=[Path("/home/ubuntu/xhs-account-isolation/data"),Path("/opt/xhs-account-isolation/data")]
    for root in roots:
        if not root.exists(): continue
        history=root / "ad-note-history" / "latest.json"
        if history.is_file(): return history
        latest=root / "ad-note-stats" / "latest.json"
        if latest.is_file(): candidates.append(latest)
        candidates.extend(root.glob("ad-note-stats/ad-note-stats-*.json"))
    existing=[path for path in candidates if path.is_file()]
    return max(existing,key=lambda path:path.stat().st_mtime) if existing else None

def build_xhs_ad_flow(fallback):
    immutable_path=latest_xhs_ad_immutable_history()
    if immutable_path:
        return build_xhs_ad_flow_from_immutable(immutable_path)
    summary_path=latest_xhs_ad_note_summary()
    if not summary_path:
        return fallback or {"generatedAt":"","date":"","accountRows":[],"ownerRows":[],"totalNoteCount":0}
    summary=json.loads(summary_path.read_text(encoding="utf-8"))
    if summary.get("schema")=="xhs-ad-history-v1":
        account_fields=("date","profile","accountName","team","noteCount","spend","opened","averageOpenCost","leads","averageLeadCost","status","statusLabel","error")
        owner_fields=("date","ownerAccountName","ownerUserId","team","noteCount","spend","opened","averageOpenCost","leads","averageLeadCost","ownerStatus","ownerStatusLabel")
        account_rows=[{key:row.get(key) for key in account_fields} for row in summary.get("accountRows",[]) if isinstance(row,dict)]
        owner_rows=[{key:row.get(key) for key in owner_fields} for row in summary.get("ownerRows",[]) if isinstance(row,dict)]
        owner_count=len({(row.get("ownerUserId") or row.get("ownerAccountName")) for row in owner_rows if row.get("ownerUserId") or row.get("ownerAccountName")})
        return {
            "generatedAt":str(summary.get("generated_at") or "")[:19].replace("T"," "),
            "date":str(summary.get("end_date") or summary.get("date") or ""),
            "periodLabel":str(summary.get("periodLabel") or ""),
            "aggregation":"DAY",
            "accountRows":account_rows,
            "ownerRows":owner_rows,
            "ownerAccountCount":owner_count,
            "totalNoteCount":int(summary.get("totalNoteCount") or 0),
            "totalSpend":round(float(summary.get("totalSpend") or 0),2),
            "totalOpened":int(summary.get("totalOpened") or 0),
            "totalLeads":int(summary.get("totalLeads") or 0),
            "unresolvedOwnerCount":int(summary.get("unresolvedOwnerCount") or 0),
        }
    report_date=str(summary.get("date") or "")
    account_rows=[]
    note_rows=[]
    for account in summary.get("accounts",[]):
        rows=account.get("rows") or []
        spend=sum(float(row.get("spend") or 0) for row in rows)
        opened=sum(int(row.get("private_message_opens") or 0) for row in rows)
        leads=sum(int(row.get("private_message_leads") or 0) for row in rows)
        status=str(account.get("status") or "")
        account_rows.append({
            "date":str(account.get("date") or report_date),
            "profile":str(account.get("profile") or ""),
            "accountName":str(account.get("account_name") or account.get("profile") or ""),
            "team":next((item["team"] for item in XHS_ACCOUNTS if item["profile"]==account.get("profile")),"管家团队"),
            "noteCount":len(rows),
            "spend":round(spend,2),
            "opened":opened,
            "averageOpenCost":round(spend/opened,2) if opened else None,
            "leads":leads,
            "averageLeadCost":round(spend/leads,2) if leads else None,
            "status":status,
            "statusLabel":"采集成功" if status=="ok" else "采集失败" if status=="error" else "未登录" if status=="not_logged_in" else "待采集",
            "error":str(account.get("error") or ""),
        })
        for row in rows:
            owner_name=str(row.get("owner_account_name") or "").strip()
            owner_status=str(row.get("owner_status") or "").strip()
            confirmed=bool(owner_name and owner_status=="confirmed")
            note_rows.append({
                "date":str(row.get("date") or account.get("date") or report_date),
                "noteId":str(row.get("note_id") or ""),
                "noteTitle":str(row.get("note_title") or ""),
                "ownerAccountName":owner_name,
                "ownerUserId":str(row.get("owner_user_id") or ""),
                "ownerSource":str(row.get("owner_source") or ""),
                "ownerStatus":"confirmed" if confirmed else "unresolved",
                "ownerStatusLabel":"已确认" if confirmed else "待确认",
                "adProfile":str(account.get("profile") or ""),
                "adAccountName":str(account.get("account_name") or account.get("profile") or ""),
                "spend":round(float(row.get("spend") or 0),2),
                "opened":int(row.get("private_message_opens") or 0),
                "leads":int(row.get("private_message_leads") or 0),
            })
    owner_groups={}
    for row in note_rows:
        owner_key=(row["date"],row["ownerUserId"] or row["ownerAccountName"] or "unresolved")
        owner=owner_groups.setdefault(owner_key,{
            "date":row["date"],"ownerAccountName":row["ownerAccountName"],"ownerUserId":row["ownerUserId"],
            "team":xhs_owner_team(row["ownerAccountName"]),
            "noteCount":0,"spend":0.0,"opened":0,"leads":0,
            "ownerStatus":"confirmed","ownerStatusLabel":"已确认",
        })
        owner["noteCount"]+=1
        owner["spend"]+=row["spend"]
        owner["opened"]+=row["opened"]
        owner["leads"]+=row["leads"]
        if row["ownerStatus"]!="confirmed":
            owner["ownerStatus"]="unresolved"
            owner["ownerStatusLabel"]="待确认"
    owner_rows=list(owner_groups.values())
    for row in owner_rows:
        row["spend"]=round(row["spend"],2)
        row["averageOpenCost"]=round(row["spend"]/row["opened"],2) if row["opened"] else None
        row["averageLeadCost"]=round(row["spend"]/row["leads"],2) if row["leads"] else None
    owner_rows.sort(key=lambda row:(row["date"],row["spend"],row["ownerAccountName"]),reverse=True)
    return {
        "generatedAt":str(summary.get("generated_at") or "")[:19].replace("T"," "),
        "date":report_date,
        "aggregation":str(summary.get("aggregation") or "DAY"),
        "accountRows":account_rows,
        "ownerRows":owner_rows,
        "totalNoteCount":len(note_rows),
        "totalSpend":round(sum(row["spend"] for row in note_rows),2),
        "totalOpened":sum(row["opened"] for row in note_rows),
        "totalLeads":sum(row["leads"] for row in note_rows),
        "unresolvedOwnerCount":sum(row["ownerStatus"]!="confirmed" for row in note_rows),
    }

def mask_xhs_email(value):
    email=str(value or "").strip()
    if "@" not in email: return "—"
    local,domain=email.rsplit("@",1)
    if local.isdigit() and len(local)>6:
        masked=local[:4]+"****"+local[-2:]
    else:
        masked=local[:4]+"****"
    return masked+"@"+domain

def latest_xhs_ad_collection():
    """Locate the newest raw per-account Aurora collection batch."""
    candidates=[]
    roots=[Path("/home/ubuntu/xhs-account-isolation/data"),Path("/opt/xhs-account-isolation/data")]
    for root in roots:
        latest=root / "ad-note-stats" / "latest.json"
        if latest.is_file(): candidates.append(latest)
    return max(candidates,key=lambda path:path.stat().st_mtime) if candidates else None

def xhs_collection_times(path):
    """Return completion time and explicit success state for every collected account."""
    if not path or not path.is_file(): return {}
    content=json.loads(path.read_text(encoding="utf-8"))
    collected_at=str(content.get("generated_at") or "")[:19].replace("T"," ")
    if not collected_at: return {}
    return {
        str(account.get("profile") or ""):{"collectedAt":collected_at,"success":str(account.get("status") or "").lower()=="ok"}
        for account in content.get("accounts",[])
        if isinstance(account,dict) and account.get("profile")
    }

def build_xhs_account_audit(fallback):
    """Derive login health and per-source collection times from the newest raw batches."""
    summary_path=latest_xhs_lead_summary()
    if not summary_path: return fallback or {}
    summary=json.loads(summary_path.read_text(encoding="utf-8"))
    end_date=summary.get("period",{}).get("end_date","")
    rows_by_profile=defaultdict(list)
    for row in summary.get("account_time_rows",[]):
        if row.get("profile") and row.get("date"):
            rows_by_profile[row["profile"]].append(row)
    content_summary_path=latest_xhs_summary()
    content_collections=xhs_collection_times(content_summary_path.with_name("latest.json") if content_summary_path else None)
    lead_collections=xhs_collection_times(summary_path.with_name("latest.json"))
    ad_collections=xhs_collection_times(latest_xhs_ad_collection())
    accounts=[]
    for account in XHS_ACCOUNTS:
        profile=account["profile"]
        ad_collection=ad_collections.get(profile,{})
        lead_collection=lead_collections.get(profile,{})
        note_collection=content_collections.get(profile,{})
        rows=sorted(rows_by_profile.get(profile,[]),key=lambda row:row.get("date",""))
        latest=rows[-1] if rows else {}
        successful=[row for row in rows if row.get("data_status")=="有数据"]
        last_success=successful[-1].get("date","") if successful else ""
        email=str(latest.get("email") or "").strip()
        xhs_id=str(latest.get("xiaohongshu_id") or "").strip()
        last_collected=latest.get("date","")
        ok=bool(email and xhs_id and last_collected and last_collected==end_date)
        if ok:
            status_label="已登录并复核"
            status_hint=f"采集至 {last_collected}"+(f"；最近有数据 {last_success}" if last_success else "")
        elif not email or not xhs_id:
            status_label="需要处理"
            status_hint="缺少登录邮箱或小红书号"
        else:
            status_label="待复核"
            status_hint=f"最近采集 {last_collected or '—'}"
        accounts.append({
            **account,
            "email":mask_xhs_email(email),
            "xiaohongshuId":xhs_id or "—",
            "lastCollectedDate":last_collected,
            "lastSuccessDate":last_success,
            "adCollectedAt":ad_collection.get("collectedAt",""),
            "adCollectedOk":bool(ad_collection.get("success")),
            "leadCollectedAt":lead_collection.get("collectedAt",""),
            "leadCollectedOk":bool(lead_collection.get("success")),
            "noteCollectedAt":note_collection.get("collectedAt",""),
            "noteCollectedOk":bool(note_collection.get("success")),
            "status":"ok" if ok else "warn",
            "statusLabel":status_label,
            "statusHint":status_hint,
        })
    ok_count=sum(1 for account in accounts if account["status"]=="ok")
    return {
        "reviewedAt":summary.get("generated_at","")[:19].replace("T"," "),
        "total":len(accounts),
        "okCount":ok_count,
        "needsAttentionCount":len(accounts)-ok_count,
        "accounts":accounts,
    }

def build_xhs_leads(fallback):
    summary_path=latest_xhs_lead_summary()
    if not summary_path:
        return attach_xhs_lead_targets(fallback or {"generatedAt":"","month":"","weeks":[],"accounts":XHS_ACCOUNTS,"dailyRows":[]})
    summary=json.loads(summary_path.read_text(encoding="utf-8"))
    end_date=datetime.strptime(summary["period"]["end_date"],"%Y-%m-%d").date()
    month_prefix=str(end_date.month)
    week_rows=[row for row in summary.get("weekly_totals",[]) if row.get("week_label","").startswith(month_prefix+"W") and row.get("week_start","")<=end_date.isoformat()]
    weeks=[{"key":row["week_label"][len(month_prefix):].lower(),"label":row["week_label"][len(month_prefix):],"start":row["week_start"],"end":row["week_end"]} for row in week_rows]
    by_profile_week={(row.get("profile"),row.get("week_label")):row for row in summary.get("account_week_rows",[])}
    accounts=[]
    for account in XHS_ACCOUNTS:
        metrics={}
        for week in weeks:
            source=by_profile_week.get((account["profile"],month_prefix+week["label"]))
            metrics[week["key"]]={
                "inbound":source.get("private_message_inbound_users") if source else None,
                "opened":source.get("private_message_opened_users") if source else None,
                "copied":source.get("personal_wechat_copy_leads") if source else None,
            }
        accounts.append({**account,"metrics":metrics})
    detail_start=(end_date-timedelta(days=20)).isoformat()
    daily_rows=[]
    for row in summary.get("account_time_rows",[]):
        row_date=row.get("date","")
        if detail_start <= row_date <= end_date.isoformat():
            daily_rows.append({
                "date":row_date,
                "week":row.get("week_label",""),
                "profile":row.get("profile",""),
                "opened":row.get("private_message_opened_users",0),
                "copied":row.get("personal_wechat_copy_leads",0),
                "status":row.get("data_status",""),
            })
    return attach_xhs_lead_targets({"generatedAt":summary.get("generated_at","")[:19].replace("T"," "),"month":f"{end_date.year:04d}-{end_date.month:02d}","weeks":weeks,"accounts":accounts,"detailStart":detail_start,"detailEnd":end_date.isoformat(),"dailyRows":daily_rows})

def extract_data(path):
    text = path.read_text(encoding="utf-8")
    match = re.search(r"const DATA\s*=\s*(\{.*?\});", text, re.S)
    if not match: raise RuntimeError(f"DATA payload missing: {path}")
    return text, json.loads(match.group(1)), match.span(1)

template, old, payload_span = extract_data(template_path)
_, current, _ = extract_data(source_path)
as_of = datetime.strptime(current["dataDate"], "%Y-%m-%d").date()

def norm(value): return str(value or "").strip().replace(" ", "")
def iso(value):
    if isinstance(value, datetime): return value.date().isoformat()
    if isinstance(value, (int,float)): return (datetime(1899,12,30)+timedelta(days=float(value))).date().isoformat()
    text=str(value or "").strip()
    return text[:10] if re.match(r"\d{4}-\d{2}-\d{2}",text) else ""
def idx(headers, *names):
    cleaned=[norm(x) for x in headers]
    for name in names:
        if norm(name) in cleaned: return cleaned.index(norm(name))
    raise RuntimeError(f"Missing column {names}")

def money(value):
    text=re.sub(r"[^0-9.\-]","",str(value or ""))
    try: return float(text) if text not in {"","-",".","-."} else 0.0
    except ValueError: return 0.0

def signing_category(source):
    source = norm(source)
    if source == "新签": return "new"
    if source == "续租": return "renewal"
    if source in {"换房", "重签"}: return "other"
    return "uncategorized"

def recent_performance():
    """Build seven-day totals by day, project, and signer from hourly exports."""
    start = as_of - timedelta(days=6)
    rows = {start + timedelta(days=i): {"newCount":0,"newRevenue":0.0,"renewalCount":0,"renewalRevenue":0.0,"otherCount":0,"actualCheckoutCount":0,"reservationCount":0,"reservationRevenue":0.0} for i in range(7)}
    projects = defaultdict(lambda:{"newCount":0,"renewalCount":0,"otherCount":0,"actualCheckoutCount":0})
    people = defaultdict(lambda:{"newCount":0,"renewalCount":0,"otherCount":0,"actualCheckoutCount":0})
    seen_signing, seen_checkout = set(), set()
    for filename in ("在租中合同.xlsx", "将搬入合同.xlsx", "已退租合同.xlsx"):
        ws = load_workbook(run_dir/filename, read_only=True, data_only=True).active
        headers = [cell.value for cell in ws[3]]
        ci, ti, si, ri = idx(headers,"合同编号"), idx(headers,"签约来源"), idx(headers,"签约时间"), idx(headers,"预退/实退")
        pi, person_i, total_rent_i = idx(headers,"所属部门"), idx(headers,"签约人"), idx(headers,"总租金")
        reason_i = idx(headers,"退租原因") if filename == "已退租合同.xlsx" else None
        for row in ws.iter_rows(min_row=4, values_only=True):
            contract = norm(row[ci])
            if not contract: continue
            sign_date, checkout_date = iso(row[si]), iso(row[ri])
            project = mapping.get(str(row[pi] or "").strip(), str(row[pi] or "").strip() or "未归类")
            person = str(row[person_i] or "").strip() or "未填写"
            if contract not in seen_signing and sign_date:
                signed = datetime.strptime(sign_date,"%Y-%m-%d").date()
                if signed in rows:
                    category = signing_category(row[ti])
                    if category != "uncategorized":
                        field = f"{category}Count"
                        rows[signed][field] += 1; projects[project][field] += 1; people[person][field] += 1
                        if category in {"new","renewal"}:
                            rows[signed][f"{category}Revenue"] += money(row[total_rent_i])
                seen_signing.add(contract)
            if filename == "已退租合同.xlsx" and contract not in seen_checkout and checkout_date and norm(row[reason_i]) != "换房清算":
                checked_out = datetime.strptime(checkout_date,"%Y-%m-%d").date()
                if checked_out in rows:
                    rows[checked_out]["actualCheckoutCount"] += 1
                    projects[project]["actualCheckoutCount"] += 1
                    people[person]["actualCheckoutCount"] += 1
                seen_checkout.add(contract)
    reservation_ws = load_workbook(run_dir/"预定合同.xlsx", read_only=True, data_only=True).active
    reservation_headers = [cell.value for cell in reservation_ws[1]]
    reservation_id_i = idx(reservation_headers,"预定ID")
    reservation_status_i = idx(reservation_headers,"状态")
    reservation_date_i = idx(reservation_headers,"录入日期")
    reservation_rent_i = idx(reservation_headers,"租金")
    seen_reservations = set()
    for row in reservation_ws.iter_rows(min_row=2, values_only=True):
        reservation_id = norm(row[reservation_id_i])
        if not reservation_id or reservation_id in seen_reservations or norm(row[reservation_status_i]) != "已付定":
            continue
        seen_reservations.add(reservation_id)
        reservation_date = iso(row[reservation_date_i])
        if not reservation_date:
            continue
        reserved = datetime.strptime(reservation_date,"%Y-%m-%d").date()
        if reserved in rows:
            rows[reserved]["reservationCount"] += 1
            rows[reserved]["reservationRevenue"] += money(row[reservation_rent_i])
    def summed(name, parts):
        return {"name":name,**{field:sum(projects[p][field] for p in parts) for field in ("newCount","renewalCount","otherCount","actualCheckoutCount")}}
    projects["北亭项目"] = summed("北亭项目",["北亭初期项目","北亭整体项目"])
    projects["小谷围项目"] = summed("小谷围项目",["北亭研寓项目","南亭研寓项目"])
    projects["小筑项目"] = summed("小筑项目",["城北小筑A","城北小筑B","城北小筑C"])
    project_rows = [{"name":name,**projects[name]} for name in base_names]
    people_rows = [{"name":name,**values} for name,values in sorted(people.items(),key=lambda item:(-sum(item[1].values()),item[0])) if sum(values.values())]
    return [{"date":day.isoformat(),**values} for day,values in rows.items()], project_rows, people_rows

def business_trend():
    """Build a continuous 30-day, newest-first series for new signings and paid reservations."""
    start=as_of-timedelta(days=29)
    rows={start+timedelta(days=i):{"newSignCount":0,"reservationCount":0} for i in range(30)}
    seen_contracts=set()
    for filename in ("在租中合同.xlsx","将搬入合同.xlsx","已退租合同.xlsx"):
        ws=load_workbook(run_dir/filename,read_only=True,data_only=True).active
        headers=[cell.value for cell in ws[3]]
        contract_i,source_i,signed_i=idx(headers,"合同编号"),idx(headers,"签约来源"),idx(headers,"签约时间")
        for row in ws.iter_rows(min_row=4,values_only=True):
            contract_id=norm(row[contract_i])
            if not contract_id or contract_id in seen_contracts: continue
            seen_contracts.add(contract_id)
            signed_date=iso(row[signed_i])
            if signing_category(row[source_i])!="new" or not signed_date: continue
            signed=datetime.strptime(signed_date,"%Y-%m-%d").date()
            if signed in rows: rows[signed]["newSignCount"]+=1
    ws=load_workbook(run_dir/"预定合同.xlsx",read_only=True,data_only=True).active
    headers=[cell.value for cell in ws[1]]
    reservation_i,status_i,created_i=idx(headers,"预定ID"),idx(headers,"状态"),idx(headers,"录入日期")
    seen_reservations=set()
    for row in ws.iter_rows(min_row=2,values_only=True):
        reservation_id=norm(row[reservation_i])
        if not reservation_id or reservation_id in seen_reservations or norm(row[status_i])!="已付定": continue
        seen_reservations.add(reservation_id)
        created_date=iso(row[created_i])
        if not created_date: continue
        created=datetime.strptime(created_date,"%Y-%m-%d").date()
        if created in rows: rows[created]["reservationCount"]+=1
    trend_rows=[{"date":day.isoformat(),**rows[day]} for day in sorted(rows,reverse=True)]
    def summary_ranges(grouping):
        ranges=[]
        for row in trend_rows:
            day=datetime.strptime(row["date"],"%Y-%m-%d").date()
            week="W1" if day.day<=7 else "W2" if day.day<=14 else "W3" if day.day<=21 else "W4" if day.day<=28 else "WE"
            period_key=f"{day.year:04d}-{day.month:02d}-{week}" if grouping=="week" else f"{day.year:04d}-{day.month:02d}"
            if not ranges or ranges[-1]["periodKey"]!=period_key:
                ranges.append({
                    "periodKey":period_key,
                    "month":f"{day.year:04d}-{day.month:02d}",
                    "week":week if grouping=="week" else "",
                    "startDate":row["date"],
                    "endDate":row["date"],
                    "dayCount":1,
                    "newSignCount":int(row["newSignCount"]),
                    "reservationCount":int(row["reservationCount"]),
                })
            else:
                ranges[-1]["endDate"]=row["date"]
                ranges[-1]["dayCount"]+=1
                ranges[-1]["newSignCount"]+=int(row["newSignCount"])
                ranges[-1]["reservationCount"]+=int(row["reservationCount"])
        for item in ranges:
            item["totalCount"]=item["newSignCount"]+item["reservationCount"]
        return ranges
    trend_ranges=summary_ranges("week")
    trend_months=summary_ranges("month")
    validation={
        "thirtyDays":len(trend_rows)==30,
        "newestFirst":all(trend_rows[i]["date"]>trend_rows[i+1]["date"] for i in range(len(trend_rows)-1)),
        "continuous":all((datetime.strptime(trend_rows[i]["date"],"%Y-%m-%d").date()-datetime.strptime(trend_rows[i+1]["date"],"%Y-%m-%d").date()).days==1 for i in range(len(trend_rows)-1)),
        "nonNegative":all(row["newSignCount"]>=0 and row["reservationCount"]>=0 for row in trend_rows),
        "weekRangeCoverage":sum(item["dayCount"] for item in trend_ranges)==len(trend_rows),
        "weekRangeTotalsMatched":sum(item["totalCount"] for item in trend_ranges)==sum(row["newSignCount"]+row["reservationCount"] for row in trend_rows),
        "monthRangeCoverage":sum(item["dayCount"] for item in trend_months)==len(trend_rows),
        "monthRangeTotalsMatched":sum(item["totalCount"] for item in trend_months)==sum(row["newSignCount"]+row["reservationCount"] for row in trend_rows),
    }
    if not all(validation.values()): raise RuntimeError(f"Business trend validation failed: {validation}")
    return {"asOfDate":as_of.isoformat(),"startDate":start.isoformat(),"endDate":as_of.isoformat(),"rows":trend_rows,"ranges":trend_ranges,"months":trend_months,"validation":validation}

def checkout_trends():
    """Build future checkout from active/move-in contracts and past actual checkout from terminated contracts."""
    past_start=as_of-timedelta(days=29)
    future_start=as_of+timedelta(days=1)
    future_end=as_of+timedelta(days=30)
    past={past_start+timedelta(days=i):0 for i in range(30)}
    past_reasons={day:{"expiryCount":0,"breachCount":0,"renewalCount":0,"otherCount":0} for day in past}
    future={future_start+timedelta(days=i):0 for i in range(30)}
    future_seen=set()
    future_source_counts=defaultdict(int)
    month_seen=set()
    month_week_counts=defaultdict(int)
    for filename in ("在租中合同.xlsx","将搬入合同.xlsx"):
        ws=load_workbook(run_dir/filename,read_only=True,data_only=True).active
        headers=[cell.value for cell in ws[3]]
        contract_i,planned_checkout_i=idx(headers,"合同编号"),idx(headers,"退租时间")
        for row in ws.iter_rows(min_row=4,values_only=True):
            contract_id=norm(row[contract_i])
            planned_checkout_date=iso(row[planned_checkout_i])
            if not contract_id or contract_id in future_seen or not planned_checkout_date: continue
            future_seen.add(contract_id)
            planned_checkout_day=datetime.strptime(planned_checkout_date,"%Y-%m-%d").date()
            if planned_checkout_day in future:
                future[planned_checkout_day]+=1
                future_source_counts[filename]+=1
            if contract_id not in month_seen and planned_checkout_date.startswith(f"{as_of.year:04d}-{as_of.month:02d}-"):
                month_seen.add(contract_id)
                day=planned_checkout_day.day
                week_key="w1" if day<=7 else "w2" if day<=14 else "w3" if day<=21 else "w4" if day<=28 else "we"
                month_week_counts[week_key]+=1
    ws=load_workbook(run_dir/"已退租合同.xlsx",read_only=True,data_only=True).active
    headers=[cell.value for cell in ws[3]]
    contract_i,actual_checkout_i,reason_i=idx(headers,"合同编号"),idx(headers,"预退/实退"),idx(headers,"退租原因")
    past_seen=set()
    for row in ws.iter_rows(min_row=4,values_only=True):
        contract_id=norm(row[contract_i])
        if not contract_id or contract_id in past_seen: continue
        past_seen.add(contract_id)
        if norm(row[reason_i])=="换房清算": continue
        actual_checkout_date=iso(row[actual_checkout_i])
        if actual_checkout_date:
            actual_checkout_day=datetime.strptime(actual_checkout_date,"%Y-%m-%d").date()
            if actual_checkout_day in past:
                past[actual_checkout_day]+=1
                reason=norm(row[reason_i])
                reason_key="expiryCount" if "到期" in reason else "breachCount" if "违约" in reason else "renewalCount" if "续租" in reason else "otherCount"
                past_reasons[actual_checkout_day][reason_key]+=1
    past_rows=[{"date":day.isoformat(),"checkoutCount":past[day]} for day in sorted(past,reverse=True)]
    past_reason_rows=[]
    for day in sorted(past_reasons,reverse=True):
        counts=past_reasons[day]
        displayed_total=counts["expiryCount"]+counts["renewalCount"]+counts["breachCount"]
        past_reason_rows.append({"date":day.isoformat(),**counts,"displayTotalCount":displayed_total,"totalCount":displayed_total+counts["otherCount"]})
    future_rows=[{"date":day.isoformat(),"checkoutCount":future[day]} for day in sorted(future)]
    def month_week(day):
        if day.day<=7: return "W1"
        if day.day<=14: return "W2"
        if day.day<=21: return "W3"
        if day.day<=28: return "W4"
        return "WE"

    def week_ranges(rows):
        ranges=[]
        for row in rows:
            day=datetime.strptime(row["date"],"%Y-%m-%d").date()
            week=month_week(day)
            period_key=f"{day.year:04d}-{day.month:02d}-{week}"
            if not ranges or ranges[-1]["periodKey"]!=period_key:
                ranges.append({
                    "index":len(ranges)+1,
                    "periodKey":period_key,
                    "month":f"{day.year:04d}-{day.month:02d}",
                    "week":week,
                    "label":f"{day.month}月 {week}",
                    "startDate":row["date"],
                    "endDate":row["date"],
                    "dayCount":1,
                    "checkoutCount":row["checkoutCount"],
                })
            else:
                ranges[-1]["endDate"]=row["date"]
                ranges[-1]["dayCount"]+=1
                ranges[-1]["checkoutCount"]+=row["checkoutCount"]
        return ranges

    def month_ranges(rows):
        ranges=[]
        for row in rows:
            day=datetime.strptime(row["date"],"%Y-%m-%d").date()
            period_key=f"{day.year:04d}-{day.month:02d}"
            if not ranges or ranges[-1]["periodKey"]!=period_key:
                ranges.append({
                    "index":len(ranges)+1,
                    "periodKey":period_key,
                    "month":period_key,
                    "label":f"{day.month}月",
                    "startDate":row["date"],
                    "endDate":row["date"],
                    "dayCount":1,
                    "checkoutCount":row["checkoutCount"],
                })
            else:
                ranges[-1]["endDate"]=row["date"]
                ranges[-1]["dayCount"]+=1
                ranges[-1]["checkoutCount"]+=row["checkoutCount"]
        return ranges

    reason_keys=("expiryCount","breachCount","renewalCount")
    def reason_summary_ranges(rows,grouping):
        ranges=[]
        for row in rows:
            day=datetime.strptime(row["date"],"%Y-%m-%d").date()
            week=month_week(day)
            group_key=f"{day.year:04d}-{day.month:02d}-{week}" if grouping=="week" else f"{day.year:04d}-{day.month:02d}"
            if not ranges or ranges[-1]["periodKey"]!=group_key:
                ranges.append({
                    "index":len(ranges)+1,
                    "periodKey":group_key,
                    "month":f"{day.year:04d}-{day.month:02d}",
                    "week":week if grouping=="week" else "",
                    "label":f"{day.month}月 {week}" if grouping=="week" else f"{day.month}月",
                    "startDate":row["date"],
                    "endDate":row["date"],
                    "dayCount":1,
                    **{key:int(row[key]) for key in reason_keys},
                    "totalCount":int(row["displayTotalCount"]),
                })
            else:
                ranges[-1]["endDate"]=row["date"]
                ranges[-1]["dayCount"]+=1
                for key in reason_keys: ranges[-1][key]+=int(row[key])
                ranges[-1]["totalCount"]+=int(row["displayTotalCount"])
        for item in ranges:
            item["renewalRate"]=item["renewalCount"]/item["totalCount"] if item["totalCount"] else 0
        return ranges

    future_ranges=week_ranges(future_rows)
    past_ranges=week_ranges(past_rows)
    future_months=month_ranges(future_rows)
    past_months=month_ranges(past_rows)
    past_reason_ranges=reason_summary_ranges(past_reason_rows,"week")
    past_reason_months=reason_summary_ranges(past_reason_rows,"month")
    occupancy_week_counts={period["key"]:sum(int(values.get(period["key"],0)) for values in checkout.values()) for period in periods}
    validation={
        "thirtyDaysEach":len(past_rows)==30 and len(future_rows)==30,
        "pastNewestFirst":all(past_rows[i]["date"]>past_rows[i+1]["date"] for i in range(len(past_rows)-1)),
        "futureNearestFirst":all(future_rows[i]["date"]<future_rows[i+1]["date"] for i in range(len(future_rows)-1)),
        "pastContinuous":all((datetime.strptime(past_rows[i]["date"],"%Y-%m-%d").date()-datetime.strptime(past_rows[i+1]["date"],"%Y-%m-%d").date()).days==1 for i in range(len(past_rows)-1)),
        "futureContinuous":all((datetime.strptime(future_rows[i+1]["date"],"%Y-%m-%d").date()-datetime.strptime(future_rows[i]["date"],"%Y-%m-%d").date()).days==1 for i in range(len(future_rows)-1)),
        "boundariesValid":past_rows[0]["date"]==as_of.isoformat() and past_rows[-1]["date"]==past_start.isoformat() and future_rows[0]["date"]==future_start.isoformat() and future_rows[-1]["date"]==future_end.isoformat(),
        "nonNegative":all(row["checkoutCount"]>=0 for row in past_rows+future_rows),
        "occupancyReconciled":all(month_week_counts[period["key"]]==occupancy_week_counts[period["key"]] for period in periods),
        "pastRangeCoverage":bool(past_ranges) and sum(item["dayCount"] for item in past_ranges)==len(past_rows) and past_ranges[0]["startDate"]==past_rows[0]["date"] and past_ranges[-1]["endDate"]==past_rows[-1]["date"] and all(item["week"] in ("W1","W2","W3","W4","WE") for item in past_ranges),
        "pastRangeTotalsMatched":sum(item["checkoutCount"] for item in past_ranges)==sum(row["checkoutCount"] for row in past_rows),
        "pastMonthCoverage":bool(past_months) and sum(item["dayCount"] for item in past_months)==len(past_rows) and past_months[0]["startDate"]==past_rows[0]["date"] and past_months[-1]["endDate"]==past_rows[-1]["date"],
        "pastMonthTotalsMatched":sum(item["checkoutCount"] for item in past_months)==sum(row["checkoutCount"] for row in past_rows),
        "pastReasonDatesMatched":[row["date"] for row in past_reason_rows]==[row["date"] for row in past_rows],
        "pastReasonDailyReconciled":all(reason_row["totalCount"]==past_row["checkoutCount"] for reason_row,past_row in zip(past_reason_rows,past_rows)),
        "pastReasonDisplayedDailyReconciled":all(row["displayTotalCount"]==row["expiryCount"]+row["renewalCount"]+row["breachCount"] and row["totalCount"]==row["displayTotalCount"]+row["otherCount"] for row in past_reason_rows),
        "pastReasonNonNegative":all(value>=0 for row in past_reason_rows for key,value in row.items() if key.endswith("Count")),
        "pastReasonRangeCoverage":sum(item["dayCount"] for item in past_reason_ranges)==len(past_reason_rows),
        "pastReasonRangeTotalsMatched":all(sum(item[key] for item in past_reason_ranges)==sum(row["displayTotalCount"] if key=="totalCount" else row[key] for row in past_reason_rows) for key in (*reason_keys,"totalCount")),
        "pastReasonMonthCoverage":sum(item["dayCount"] for item in past_reason_months)==len(past_reason_rows),
        "pastReasonMonthTotalsMatched":all(sum(item[key] for item in past_reason_months)==sum(row["displayTotalCount"] if key=="totalCount" else row[key] for row in past_reason_rows) for key in (*reason_keys,"totalCount")),
        "futureRangeCoverage":bool(future_ranges) and sum(item["dayCount"] for item in future_ranges)==len(future_rows) and future_ranges[0]["startDate"]==future_rows[0]["date"] and future_ranges[-1]["endDate"]==future_rows[-1]["date"] and all(item["week"] in ("W1","W2","W3","W4","WE") for item in future_ranges),
        "futureRangeTotalsMatched":sum(item["checkoutCount"] for item in future_ranges)==sum(row["checkoutCount"] for row in future_rows),
        "futureMonthCoverage":bool(future_months) and sum(item["dayCount"] for item in future_months)==len(future_rows) and future_months[0]["startDate"]==future_rows[0]["date"] and future_months[-1]["endDate"]==future_rows[-1]["date"],
        "futureMonthTotalsMatched":sum(item["checkoutCount"] for item in future_months)==sum(row["checkoutCount"] for row in future_rows),
    }
    if not all(validation.values()): raise RuntimeError(f"Checkout trend validation failed: {validation}")
    return {
        "asOfDate":as_of.isoformat(),
        "past":{"startDate":past_start.isoformat(),"endDate":as_of.isoformat(),"sourceFiles":["已退租合同.xlsx"],"dateField":"预退/实退","reasonField":"退租原因","reasonCategories":["到期","违约","续租","其他"],"displayedReasonCategories":["到期","续租","违约"],"rows":past_rows,"reasonRows":past_reason_rows,"ranges":past_ranges,"months":past_months,"reasonRanges":past_reason_ranges,"reasonMonths":past_reason_months},
        "future":{"startDate":future_start.isoformat(),"endDate":future_end.isoformat(),"sourceFiles":["在租中合同.xlsx","将搬入合同.xlsx"],"dateField":"退租时间","sourceCounts":dict(future_source_counts),"rows":future_rows,"ranges":future_ranges,"months":future_months},
        "occupancyWeekCounts":occupancy_week_counts,
        "validation":validation,
    }

def overview_contract_activity(recent_rows,current_month,monthly_details):
    """Summarize deduplicated contract events for today, yesterday, current week, and current month."""
    by_date={row["date"]:row for row in recent_rows}
    week_start=as_of-timedelta(days=as_of.weekday())
    fields={
        "newSign":"newCount",
        "newSignRevenue":"newRevenue",
        "reservation":"reservationCount",
        "reservationRevenue":"reservationRevenue",
        "renewal":"renewalCount",
        "renewalRevenue":"renewalRevenue",
        "actualCheckout":"actualCheckoutCount",
    }
    def period(key,label,start,end):
        selected=[row for row in recent_rows if start.isoformat()<=row["date"]<=end.isoformat()]
        return {
            "key":key,
            "label":label,
            "startDate":start.isoformat(),
            "endDate":end.isoformat(),
            "metrics":{name:round(sum(float(row.get(source) or 0) for row in selected),2) if name.endswith("Revenue") else sum(int(row.get(source) or 0) for row in selected) for name,source in fields.items()},
        }
    def metric_values(row):
        return {name:round(float(row.get(source) or 0),2) if name.endswith("Revenue") else int(row.get(source) or 0) for name,source in fields.items()}
    yesterday=as_of-timedelta(days=1)
    month_start=as_of.replace(day=1)
    reservation_ws=load_workbook(run_dir/"预定合同.xlsx",read_only=True,data_only=True).active
    reservation_headers=[cell.value for cell in reservation_ws[1]]
    reservation_id_i,status_i,created_i,reservation_rent_i=idx(reservation_headers,"预定ID"),idx(reservation_headers,"状态"),idx(reservation_headers,"录入日期"),idx(reservation_headers,"租金")
    reservation_address_i,reservation_building_i=idx(reservation_headers,"地址"),idx(reservation_headers,"小区/公寓")
    reservation_start_i,reservation_end_i,reservation_signer_i=idx(reservation_headers,"合同开始"),idx(reservation_headers,"合同结束"),idx(reservation_headers,"预定办理人")
    seen_month_reservations=set()
    month_reservations=0
    month_reservation_revenue=0.0
    reservation_details=[]
    for row in reservation_ws.iter_rows(min_row=2,values_only=True):
        reservation_id=norm(row[reservation_id_i])
        if not reservation_id or reservation_id in seen_month_reservations or norm(row[status_i])!="已付定": continue
        seen_month_reservations.add(reservation_id)
        created_date=iso(row[created_i])
        if created_date and month_start.isoformat()<=created_date<=as_of.isoformat():
            month_reservations+=1
            month_reservation_revenue+=money(row[reservation_rent_i])
            address=str(row[reservation_address_i] or "").strip()
            room_numbers=re.findall(r"\d+",address)
            reservation_details.append({
                "date":created_date,
                "building":str(row[reservation_building_i] or "").strip() or address or "—",
                "roomNo":room_numbers[-1] if room_numbers else "—",
                "leaseStart":iso(row[reservation_start_i]) or "—",
                "leaseEnd":iso(row[reservation_end_i]) or "—",
                "rent":round(money(row[reservation_rent_i]),2),
                "signer":str(row[reservation_signer_i] or "").strip() or "—",
            })
    periods=[
        period("today","今天",as_of,as_of),
        period("yesterday","昨天",yesterday,yesterday),
        period("week","本周",week_start,as_of),
        {
            "key":"month","label":"本月","startDate":month_start.isoformat(),"endDate":as_of.isoformat(),
            "metrics":{
                "newSign":int(current_month.get("newCount") or 0),
                "newSignRevenue":round(float(current_month.get("newRevenue") or 0),2),
                "reservation":month_reservations,
                "reservationRevenue":round(month_reservation_revenue,2),
                "renewal":int(current_month.get("renewalCount") or 0),
                "renewalRevenue":round(float(current_month.get("renewalRevenue") or 0),2),
                "actualCheckout":int(current_month.get("actualCheckoutCount") or 0),
            },
        },
    ]
    source_details={
        "newSign":monthly_details.get("new",[]),
        "reservation":reservation_details,
        "renewal":monthly_details.get("renewal",[]),
        "actualCheckout":monthly_details.get("actualCheckout",[]),
    }
    for item in periods:
        start_date,end_date=item["startDate"],item["endDate"]
        item["details"]={
            key:[
                {field:row.get(field,"—") for field in ("building","roomNo","leaseStart","leaseEnd","rent","signer")}
                for row in sorted(rows,key=lambda value:(value.get("date","") ,value.get("building","") ,value.get("roomNo","")),reverse=True)
                if start_date<=row.get("date","")<=end_date
            ]
            for key,rows in source_details.items()
        }
    validation={
        "todayMatched":periods[0]["metrics"]==metric_values(by_date.get(as_of.isoformat(),{})),
        "yesterdayMatched":periods[1]["metrics"]==metric_values(by_date.get(yesterday.isoformat(),{})),
        "weekRangeValid":week_start<=as_of and (as_of-week_start).days<7,
        "monthRangeValid":month_start<=as_of and month_start.day==1,
        "monthCoreMatched":periods[3]["metrics"]=={
            "newSign":int(current_month.get("newCount") or 0),
            "newSignRevenue":round(float(current_month.get("newRevenue") or 0),2),
            "reservation":month_reservations,
            "reservationRevenue":round(month_reservation_revenue,2),
            "renewal":int(current_month.get("renewalCount") or 0),
            "renewalRevenue":round(float(current_month.get("renewalRevenue") or 0),2),
            "actualCheckout":int(current_month.get("actualCheckoutCount") or 0),
        },
        "detailCountsMatched":all(len(item["details"][key])==int(item["metrics"][key]) for item in periods for key in ("newSign","reservation","renewal","actualCheckout")),
        "nonNegative":all(value>=0 for item in periods for value in item["metrics"].values()),
    }
    if not all(validation.values()): raise RuntimeError(f"Overview contract activity validation failed: {periods}")
    return {"asOfDate":as_of.isoformat(),"weekStart":week_start.isoformat(),"periods":periods,"validation":validation}

def monthly_contract_details():
    """Build current-month new, renewal, other, and actual-checkout detail rows."""
    month_prefix = f"{as_of.year:04d}-{as_of.month:02d}"
    details = {"new": [], "renewal": [], "other": [], "actualCheckout": []}

    def detail_row(row, indexes, event_date):
        def value(name):
            raw = row[indexes[name]]
            return str(raw).strip() if raw is not None and str(raw).strip() else "—"
        room_match = re.search(r"\d+", value("门牌号"))
        return {
            "date": event_date,
            "signSource": value("签约来源"),
            "building": value("小区/公寓"),
            "roomNo": room_match.group(0) if room_match else "—",
            "customerName": value("租客姓名"),
            "leaseStart": iso(row[indexes["起租时间"]]) or "—",
            "leaseEnd": iso(row[indexes["退租时间"]]) or "—",
            "leasePeriod": value("租期时长"),
            "rent": round(money(row[indexes["租金单价"]]),2),
            "signer": value("签约人"),
            "contractId": norm(row[indexes["合同编号"]]),
        }

    seen_signing = set()
    for filename in ("在租中合同.xlsx", "将搬入合同.xlsx", "已退租合同.xlsx"):
        ws = load_workbook(run_dir/filename, read_only=True, data_only=True).active
        headers = [cell.value for cell in ws[3]]
        names = ("合同编号", "签约来源", "签约时间", "小区/公寓", "门牌号", "租客姓名", "起租时间", "退租时间", "租期时长", "租金单价", "签约人")
        indexes = {name: idx(headers, name) for name in names}
        for row in ws.iter_rows(min_row=4, values_only=True):
            contract_id = norm(row[indexes["合同编号"]])
            if not contract_id or contract_id in seen_signing:
                continue
            seen_signing.add(contract_id)
            sign_date = iso(row[indexes["签约时间"]])
            if not sign_date.startswith(month_prefix):
                continue
            key = signing_category(row[indexes["签约来源"]])
            if key != "uncategorized":
                details[key].append(detail_row(row, indexes, sign_date))

    ws = load_workbook(run_dir/"已退租合同.xlsx", read_only=True, data_only=True).active
    headers = [cell.value for cell in ws[3]]
    names = ("合同编号", "签约来源", "预退/实退", "退租原因", "小区/公寓", "门牌号", "租客姓名", "起租时间", "退租时间", "租期时长", "租金单价", "签约人")
    indexes = {name: idx(headers, name) for name in names}
    seen_checkout = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        contract_id = norm(row[indexes["合同编号"]])
        if not contract_id or contract_id in seen_checkout:
            continue
        seen_checkout.add(contract_id)
        checkout_date = iso(row[indexes["预退/实退"]])
        if not checkout_date.startswith(month_prefix) or norm(row[indexes["退租原因"]]) == "换房清算":
            continue
        details["actualCheckout"].append(detail_row(row, indexes, checkout_date))

    for rows in details.values():
        rows.sort(key=lambda row: (row["date"], row["contractId"]), reverse=True)
    return {"month": month_prefix, **details}

building_rows={row["name"]:row for row in current["buildingData"]}

def build_overview_new():
    """Independent room-level overview; does not change existing dashboard metrics."""
    def room_keys(room_id,address):
        keys=set()
        if norm(room_id): keys.add("id:"+norm(room_id))
        if norm(address): keys.add("address:"+norm(address))
        return keys

    def contract_room_keys(filename,header_row,required_status,status_names,address_names):
        contract_ws=load_workbook(run_dir/filename,read_only=True,data_only=True).active
        contract_headers=[cell.value for cell in contract_ws[header_row]]
        room_i=idx(contract_headers,"房源ID")
        address_i=idx(contract_headers,*address_names)
        status_i=idx(contract_headers,*status_names)
        keys=set()
        for row in contract_ws.iter_rows(min_row=header_row+1,values_only=True):
            if norm(row[status_i])==required_status:
                keys.update(room_keys(row[room_i],row[address_i]))
        return keys

    preorder_keys=contract_room_keys("预定合同.xlsx",1,"已付定",("状态",),("地址","房源地址"))
    moving_keys=contract_room_keys("将搬入合同.xlsx",3,"将搬入",("合同状态","状态"),("房源地址","地址"))
    ws=load_workbook(run_dir/"房源详情.xlsx",read_only=True,data_only=True).active
    headers=[cell.value for cell in ws[3]]
    room_i=idx(headers,"房源ID")
    address_i=idx(headers,"房源地址","地址")
    status_i=idx(headers,"状态")
    lock_i=idx(headers,"锁房备注")
    result={"totalRooms":0,"occupiedCount":0,"rentableCount":0,"unavailableCount":0,"lockedCount":0,"preorderCount":0,"moveInCount":0,"otherUnavailableCount":0,"shortRentCount":0,"comprehensiveCount":0,"sourceOtherStatusCount":0}
    status_counts=defaultdict(int)
    occupied_rooms=set()
    preorder_rooms=set()
    moving_rooms=set()
    short_rent_rooms=set()
    overlaps={"lockedAndPreorder":0,"lockedAndMoveIn":0,"preorderAndMoveIn":0}
    known_states={"已出租","在租中","空房可租","空房不可租"}
    for row in ws.iter_rows(min_row=4,values_only=True):
        if not any(value not in (None,"") for value in row): continue
        result["totalRooms"]+=1
        state=norm(row[status_i])
        remark=norm(row[lock_i])
        keys=room_keys(row[room_i],row[address_i])
        room_identity=next((key for key in keys if key.startswith("id:")),next(iter(keys),f"row:{result['totalRooms']}"))
        status_counts[state or "空白状态"]+=1
        if state in {"已出租","在租中"}:
            result["occupiedCount"]+=1
            occupied_rooms.add(room_identity)
            continue
        if state=="空房可租":
            result["rentableCount"]+=1
            continue
        result["unavailableCount"]+=1
        if state not in known_states: result["sourceOtherStatusCount"]+=1
        has_lock=bool(remark)
        has_preorder=bool(keys & preorder_keys)
        has_moving=bool(keys & moving_keys)
        if has_lock and has_preorder: overlaps["lockedAndPreorder"]+=1
        if has_lock and has_moving: overlaps["lockedAndMoveIn"]+=1
        if has_preorder and has_moving: overlaps["preorderAndMoveIn"]+=1
        if has_preorder:
            result["preorderCount"]+=1
            preorder_rooms.add(room_identity)
        elif has_moving:
            result["moveInCount"]+=1
            moving_rooms.add(room_identity)
        else:
            result["lockedCount"]+=1
            if "短租" in remark: short_rent_rooms.add(room_identity)
    base_comprehensive_rooms=occupied_rooms|preorder_rooms|moving_rooms
    comprehensive_rooms=base_comprehensive_rooms|short_rent_rooms
    result["shortRentCount"]=len(short_rent_rooms)
    result["shortRentOverlapCount"]=len(short_rent_rooms & base_comprehensive_rooms)
    result["comprehensiveCount"]=len(comprehensive_rooms)
    result["comprehensiveRate"]=result["comprehensiveCount"]/result["totalRooms"] if result["totalRooms"] else 0
    result["statusCounts"]=[{"status":name,"count":count} for name,count in sorted(status_counts.items())]
    result["overlapsResolved"]=overlaps
    result["priority"]=["预定","将搬入","锁房"]
    result["validation"]={
        "baseReconciled":result["occupiedCount"]+result["rentableCount"]+result["unavailableCount"]==result["totalRooms"],
        "unavailableReconciled":result["lockedCount"]+result["preorderCount"]+result["moveInCount"]+result["otherUnavailableCount"]==result["unavailableCount"],
        "shortRentWithinLocked":result["shortRentCount"]<=result["lockedCount"],
        "comprehensiveBounded":0<=result["comprehensiveCount"]<=result["totalRooms"],
    }
    if not all(result["validation"].values()): raise RuntimeError(f"Overview-new reconciliation failed: {result}")
    return result

def apply_house_monitoring():
    """Use 房源详情 as the authoritative source for locks and vacancy availability."""
    def contract_room_ids(filename, header_row, required_status, *status_names):
        contract_ws=load_workbook(run_dir/filename,read_only=True,data_only=True).active
        contract_headers=[cell.value for cell in contract_ws[header_row]]
        room_i=idx(contract_headers,"房源ID")
        contract_status_i=idx(contract_headers,*status_names)
        return {
            norm(row[room_i])
            for row in contract_ws.iter_rows(min_row=header_row+1,values_only=True)
            if norm(row[room_i]) and norm(row[contract_status_i]) == required_status
        }
    preorder_room_ids=contract_room_ids("预定合同.xlsx",1,"已付定","状态")
    moving_room_ids=contract_room_ids("将搬入合同.xlsx",3,"将搬入","合同状态","状态")
    ws=load_workbook(run_dir/"房源详情.xlsx",read_only=True,data_only=True).active
    headers=[cell.value for cell in ws[3]]
    building_i, room_i, lock_i, status_i=idx(headers,"小区/公寓"),idx(headers,"房源ID"),idx(headers,"锁房备注"),idx(headers,"状态")
    locks, rentable, unrentable=defaultdict(int),defaultdict(int),defaultdict(int)
    mapped_unknown=defaultdict(int)
    occupied, preordered, moving_in, unknown_status=defaultdict(int),defaultdict(int),defaultdict(int),defaultdict(int)
    for row in ws.iter_rows(min_row=4,values_only=True):
        building=str(row[building_i] or "").strip()
        room_id=norm(row[room_i])
        state=norm(row[status_i])
        remark=norm(row[lock_i])
        monitored=bool(remark) or state in {"已出租","在租中","空房可租","空房不可租","未知状态"}
        if monitored and not building: raise RuntimeError("Monitored room is missing building name")
        if remark: locks[building]+=1
        if state in {"已出租","在租中"}: occupied[building]+=1
        if state == "空房可租": rentable[building]+=1
        if state in {"空房不可租","未知状态"}:
            unrentable[building]+=1
            if state == "未知状态": mapped_unknown[building]+=1
            if room_id in moving_room_ids or any(word in remark for word in ("将搬入","待搬入")): moving_in[building]+=1
            elif room_id in preorder_room_ids or any(word in remark for word in ("已预订","已预定","预订","预定")): preordered[building]+=1
        if state not in {"已出租","在租中","空房可租","空房不可租","未知状态"}: unknown_status[building]+=1
    unknown=sorted((set(locks)|set(rentable)|set(unrentable)|set(occupied)|set(preordered)|set(moving_in)|set(mapped_unknown)|set(unknown_status))-set(building_rows))
    if unknown: raise RuntimeError(f"Locked rooms contain unknown buildings: {unknown}")
    for name,row in building_rows.items():
        locked=locks[name]
        row["lockCount"]=locked
        row["rentableVacancyCount"]=rentable[name]
        row["unrentableVacancyCount"]=unrentable[name]
        row["vacancyCount"]=rentable[name]+unrentable[name]
        row["occupiedCount"]=occupied[name]
        row["preorderCount"]=preordered[name]
        row["moveInCount"]=moving_in[name]
        row["qualifyingUnavailableCount"]=preordered[name]+moving_in[name]
        rooms=int(row.get("rooms",0) or 0)
        row["lockRate"]=locked/rooms if rooms else 0
        row["occupancyRate"]=row["occupiedCount"]/rooms if rooms else 0
        row["vacancyRate"]=row["vacancyCount"]/rooms if rooms else 0
        row["comprehensiveCount"]=row["occupiedCount"]+row["qualifyingUnavailableCount"]
        row["comprehensiveRate"]=row["comprehensiveCount"]/rooms if rooms else 0
    return {"lock":sum(locks.values()),"rentable":sum(rentable.values()),"unrentable":sum(unrentable.values()),"occupied":sum(occupied.values()),"preordered":sum(preordered.values()),"movingIn":sum(moving_in.values()),"mappedUnknown":sum(mapped_unknown.values()),"unknownStatus":sum(unknown_status.values())}

house_monitor=apply_house_monitoring()
lock_total=house_monitor["lock"]
periods=[
    {"key":"w1","label":f"{as_of.month}W1","range":f"{as_of.month}月1日至7日"},
    {"key":"w2","label":f"{as_of.month}W2","range":f"{as_of.month}月8日至14日"},
    {"key":"w3","label":f"{as_of.month}W3","range":f"{as_of.month}月15日至21日"},
    {"key":"w4","label":f"{as_of.month}W4","range":f"{as_of.month}月22日至28日"},
    {"key":"we","label":f"{as_of.month}WE","range":f"{as_of.month}月29日至月末"},
]
checkout=defaultdict(lambda:defaultdict(int))
seen=set()
for filename in ("在租中合同.xlsx","将搬入合同.xlsx"):
    ws=load_workbook(run_dir/filename,read_only=True,data_only=True).active
    headers=[cell.value for cell in ws[3]]
    ci,bi,di=idx(headers,"合同编号"),idx(headers,"小区/公寓"),idx(headers,"退租时间")
    for row in ws.iter_rows(min_row=4,values_only=True):
        contract=norm(row[ci]); date=iso(row[di]); building=str(row[bi] or "").strip()
        if not contract or contract in seen or not date.startswith(f"{as_of.year:04d}-{as_of.month:02d}-"): continue
        seen.add(contract); day=int(date[8:10]); key="w1" if day<=7 else "w2" if day<=14 else "w3" if day<=21 else "w4" if day<=28 else "we"
        checkout[building][key]+=1

for name,row in building_rows.items():
    row["checkout"]={p["key"]:{"count":checkout[name][p["key"]],"rate":checkout[name][p["key"]]/row["rooms"] if row["rooms"] else 0} for p in periods}

def aggregate(name, names):
    rows=[building_rows[n] for n in names if n in building_rows]
    result={"name":name}
    for field in ("rooms","comprehensiveCount","occupiedCount","vacancyCount","lockCount","rentableVacancyCount","unrentableVacancyCount","preorderCount","moveInCount","qualifyingUnavailableCount"):
        result[field]=sum(float(r.get(field,0) or 0) for r in rows)
        if result[field].is_integer(): result[field]=int(result[field])
    rooms=result["rooms"]
    for count,rate in (("comprehensiveCount","comprehensiveRate"),("occupiedCount","occupancyRate"),("vacancyCount","vacancyRate"),("lockCount","lockRate")):
        result[rate]=result[count]/rooms if rooms else 0
    result["checkout"]={p["key"]:{"count":sum(r["checkout"][p["key"]]["count"] for r in rows),"rate":0} for p in periods}
    for p in periods: result["checkout"][p["key"]]["rate"]=result["checkout"][p["key"]]["count"]/rooms if rooms else 0
    return result

all_names=list(building_rows)
bt_initial=[n for n in all_names if re.fullmatch(r"北亭0?[1-3]座",n)]
bt_whole=[n for n in all_names if re.fullmatch(r"北亭(?:0?[5-9]|1[0-5])座",n)]
bt_research=[n for n in all_names if n.startswith("北亭研寓")]
nt_research=[n for n in all_names if re.fullmatch(r"南亭(?:20|2[2-8])座",n)]
self_owned=[n for n in all_names if re.fullmatch(r"南亭(?:0?[1-9]|1[0-4])座",n)]
small_a=[n for n in all_names if n=="城北小筑01座"]
small_b=[n for n in all_names if n in {"城北小筑02座","城北小筑06座","城北小筑07座","城北小筑08座","城北小筑11座","城北小筑12座","城南小筑01座"}]
small_c=[n for n in all_names if n in {"城北小筑03座","城北小筑05座","城北小筑10座"}]

project_specs=[
 ("全部房源汇总",all_names),("北亭项目",bt_initial+bt_whole),("北亭初期项目",bt_initial),("北亭整体项目",bt_whole),
 ("小谷围项目",bt_research+nt_research),("北亭研寓项目",bt_research),("南亭研寓项目",nt_research),("公司自持项目",self_owned),
 ("小筑项目",small_a+small_b+small_c),("城北小筑A",small_a),("城北小筑B",small_b),("城北小筑C",small_c),
 ("南亭15座独立项目",["南亭15座"]),("南亭18座独立项目",["南亭18座"]),("南亭21座独立项目",["南亭21座"]),("城北小筑09座独立项目",["城北小筑09座"]),
]
project_data=[aggregate(name,names) for name,names in project_specs]
project_data += [aggregate("南亭区域汇总",[n for n in all_names if "南" in n]),aggregate("北亭区域汇总",[n for n in all_names if "北" in n])]

def build_ziyin_occupancy():
    """Future contracts and month-specific occupancy used only by 入住率（紫茵）."""
    current_month=f"{as_of.year:04d}-{as_of.month:02d}"
    next_start=(as_of.replace(day=28)+timedelta(days=4)).replace(day=1)
    next_month=f"{next_start.year:04d}-{next_start.month:02d}"
    scope_buildings={
        "南亭区域":{n for n in all_names if "南" in n and "小筑" not in n},
        "北亭区域":{n for n in all_names if "北" in n and "小筑" not in n},
        "小筑项目":set(small_a+small_b+small_c),
    }
    future_keys=("currentMoveIn","currentPreorder","nextMoveIn","nextPreorder")
    result={name:{key:{"count":0,"occupiedOverlap":0,"_rooms":[]} for key in future_keys} for name in scope_buildings}
    active_ws=load_workbook(run_dir/"在租中合同.xlsx",read_only=True,data_only=True).active
    active_headers=[cell.value for cell in active_ws[3]]
    active_room_i,active_address_i=idx(active_headers,"房源ID"),idx(active_headers,"房源地址")
    def room_keys(room_id,address):
        keys=set()
        if norm(room_id): keys.add("id:"+norm(room_id))
        if norm(address): keys.add("address:"+norm(address))
        return keys
    active_keys=set()
    for row in active_ws.iter_rows(min_row=4,values_only=True): active_keys.update(room_keys(row[active_room_i],row[active_address_i]))
    specs=[
        ("将搬入合同.xlsx",3,4,"合同编号","合同状态","将搬入","起租时间","小区/公寓","房源ID","房源地址","MoveIn"),
        ("预定合同.xlsx",1,2,"预定ID","状态","已付定","合同开始","小区/公寓","房源ID","地址","Preorder"),
    ]
    for filename,header_row,data_row,id_name,status_name,status_value,date_name,building_name,room_name,address_name,suffix in specs:
        ws=load_workbook(run_dir/filename,read_only=True,data_only=True).active
        headers=[cell.value for cell in ws[header_row]]
        unique_i,status_i,date_i,building_i,room_i,address_i=(idx(headers,id_name),idx(headers,status_name),idx(headers,date_name),idx(headers,building_name),idx(headers,room_name),idx(headers,address_name))
        seen=set()
        for row_number,row in enumerate(ws.iter_rows(min_row=data_row,values_only=True),start=data_row):
            if norm(row[status_i])!=status_value: continue
            unique_key=norm(row[unique_i]) or f"{filename}:{row_number}"
            if unique_key in seen: continue
            seen.add(unique_key)
            date_value=iso(row[date_i])
            month=date_value[:7]
            if month not in {current_month,next_month}: continue
            building=str(row[building_i] or "").strip()
            scope=next((name for name,names in scope_buildings.items() if building in names),None)
            if not scope: continue
            period="current" if month==current_month else "next"
            key=period+suffix
            keys=room_keys(row[room_i],row[address_i])
            result[scope][key]["count"]+=1
            if keys & active_keys: result[scope][key]["occupiedOverlap"]+=1
            if keys: result[scope][key]["_rooms"].append(keys)

    def unique_non_active_rooms(items):
        groups=[]
        for keys in items:
            if keys & active_keys: continue
            merged=set(keys)
            remaining=[]
            for group in groups:
                if merged & group: merged.update(group)
                else: remaining.append(group)
            remaining.append(merged)
            groups=remaining
        return len(groups)

    rows=[]
    base_by_name={name:aggregate(name,names) for name,names in scope_buildings.items()}
    for name,values in result.items():
        for item in values.values():
            if item["occupiedOverlap"]>item["count"]: raise RuntimeError("Ziyin overlap exceeds contract count")
        base=base_by_name[name]
        current_rooms=values["currentMoveIn"]["_rooms"]+values["currentPreorder"]["_rooms"]
        next_rooms=current_rooms+values["nextMoveIn"]["_rooms"]+values["nextPreorder"]["_rooms"]
        current_count=min(base["rooms"],base["occupiedCount"]+unique_non_active_rooms(current_rooms))
        next_count=min(base["rooms"],base["occupiedCount"]+unique_non_active_rooms(next_rooms))
        cleaned={key:{k:v for k,v in item.items() if k!="_rooms"} for key,item in values.items()}
        rows.append({
            "name":name,**cleaned,
            "currentComprehensiveCount":current_count,
            "currentComprehensiveRate":current_count/base["rooms"] if base["rooms"] else 0,
            "nextComprehensiveCount":next_count,
            "nextComprehensiveRate":next_count/base["rooms"] if base["rooms"] else 0,
        })

    summary_base=aggregate("汇总",set().union(*scope_buildings.values()))
    summary={**summary_base}
    for key in future_keys:
        summary[key]={field:sum(row[key][field] for row in rows) for field in ("count","occupiedOverlap")}
    summary["currentComprehensiveCount"]=sum(row["currentComprehensiveCount"] for row in rows)
    summary["nextComprehensiveCount"]=sum(row["nextComprehensiveCount"] for row in rows)
    summary["currentComprehensiveRate"]=summary["currentComprehensiveCount"]/summary["rooms"] if summary["rooms"] else 0
    summary["nextComprehensiveRate"]=summary["nextComprehensiveCount"]/summary["rooms"] if summary["rooms"] else 0
    if any(row[key]>base_by_name[row["name"]]["rooms"] for row in rows for key in ("currentComprehensiveCount","nextComprehensiveCount")):
        raise RuntimeError("Ziyin comprehensive count exceeds room count")
    return {"asOfDate":as_of.isoformat(),"currentMonth":current_month,"nextMonth":next_month,"rows":rows,"summary":summary}

ziyin_occupancy=build_ziyin_occupancy()

mapping={
 "整体项目-北亭初期框架":"北亭初期项目","整体项目-北亭整体框架":"北亭整体项目","整体项目-北亭研寓框架":"北亭研寓项目",
 "整体项目-南亭研寓框架":"南亭研寓项目","整体项目-自持物业框架":"公司自持项目","整体项目-城北小筑A":"城北小筑A",
 "整体项目-城北小筑B":"城北小筑B","整体项目-城北小筑C":"城北小筑C","独立项目-南亭15座":"南亭15座独立项目",
 "独立项目-南亭18座":"南亭18座独立项目","独立项目-南亭21座":"南亭21座独立项目","独立项目-城北小筑01座":"城北小筑09座独立项目",
}
cs=current["contractStats"]
monthly={mapping.get(r["name"],r["name"]):{**r,"name":mapping.get(r["name"],r["name"])} for r in cs.get("projectMonthly",[])}
def sum_monthly(name,parts):
    return {"name":name,**{k:sum(monthly.get(p,{}).get(k,0) for p in parts) for k in ("newCount","renewalCount","otherCount","actualCheckoutCount","checkoutRenewalCount")}}
monthly["北亭项目"]=sum_monthly("北亭项目",["北亭初期项目","北亭整体项目"])
monthly["小谷围项目"]=sum_monthly("小谷围项目",["北亭研寓项目","南亭研寓项目"])
monthly["小筑项目"]=sum_monthly("小筑项目",["城北小筑A","城北小筑B","城北小筑C"])
base_names=[name for name,_ in project_specs[1:]]
contract_stats={**old["contractStats"],**cs,"asOfDate":current["dataDate"],"projectMonthly":[monthly.get(n,{"name":n,"newCount":0,"renewalCount":0,"otherCount":0,"actualCheckoutCount":0,"checkoutRenewalCount":0}) for n in base_names]}
contract_stats["recentPerformance"], contract_stats["recentProjectPerformance"], contract_stats["recentPeoplePerformance"] = recent_performance()
contract_stats["monthlyDetails"] = monthly_contract_details()
detail_counts = [len(contract_stats["monthlyDetails"][key]) for key in ("new", "renewal", "other", "actualCheckout")]
summary_counts = [int(contract_stats["currentMonth"].get(key, 0) or 0) for key in ("newCount", "renewalCount", "otherCount", "actualCheckoutCount")]
if detail_counts != summary_counts:
    raise RuntimeError(f"Monthly contract details do not match summary: details={detail_counts}, summary={summary_counts}")
for period_name in ("currentMonth","previousMonth"):
    period=contract_stats[period_name]
    if int(period.get("actualCheckoutCount") or 0)!=int(period.get("checkoutActualDepartureCount") or 0)+int(period.get("checkoutRenewalCount") or 0):
        raise RuntimeError(f"Checkout breakdown does not reconcile: {period_name}")
contract_stats.setdefault("projectMonthlyUnmapped",{"newCount":0,"renewalCount":0,"otherCount":0,"actualCheckoutCount":0,"checkoutRenewalCount":0})
contract_stats.setdefault("totals",{})
contract_stats["uniqueContracts"]=sum(1 for f in ("在租中合同.xlsx","将搬入合同.xlsx","已退租合同.xlsx") for _ in load_workbook(run_dir/f,read_only=True).active.iter_rows(min_row=4,values_only=True))

overview_new=build_overview_new()
overview_new["contractActivity"]=overview_contract_activity(contract_stats["recentPerformance"],contract_stats["currentMonth"],contract_stats["monthlyDetails"])
payload={"dataDate":current["dataDate"],"generatedDate":datetime.now().astimezone().strftime("%Y-%m-%d %H:%M"),"projectData":project_data,"buildingData":list(building_rows.values()),"contractStats":contract_stats,"baseProjectNames":base_names,"checkoutPeriods":periods,"overviewNew":overview_new,"businessTrend":business_trend(),"checkoutTrends":checkout_trends(),"ziyinOccupancy":ziyin_occupancy,"xhsAccountAudit":build_xhs_account_audit(old.get("xhsAccountAudit")),"xhsContent":build_xhs_content(old.get("xhsContent")),"xhsNotePublished":build_xhs_note_published(old.get("xhsNotePublished")),"xhsLeads":build_xhs_leads(old.get("xhsLeads")),"xhsAdFlow":build_xhs_ad_flow(old.get("xhsAdFlow")),"customerData":build_customer_data(old.get("customerData")),"meterManagement":build_meter_management(old.get("meterManagement"))}
rendered=template[:payload_span[0]]+json.dumps(payload,ensure_ascii=False,separators=(",",":"))+template[payload_span[1]:]
required=['class="nav desktop-nav"','data-desktop-module="xiaohongshu"','data-desktop-module="yuxiaor"','data-desktop-menu="xiaohongshu"','data-desktop-menu="yuxiaor"','data-dashboard-view="operations-brief"','data-dashboard-view="overview"','data-dashboard-view="performance"','data-dashboard-view="occupancy"','data-dashboard-view="occupancy-ziyin"','id="occupancy-ziyin"','ziyin-project-table','function renderZiyinOccupancy()','"ziyinOccupancy"','occupiedOverlap','class="mobile-nav-shell"','data-mobile-menu="primary"','data-mobile-module="xiaohongshu"','data-mobile-module="yuxiaor"','data-mobile-menu="xiaohongshu"','data-mobile-menu="yuxiaor"','5%以下绿色','brief-daily-table','brief-project-table','brief-person-table','id="xhs-account"','xhs-account-table','xhs-account-updated','xhs-account-status-list','adCollectedAt','adCollectedOk','leadCollectedAt','leadCollectedOk','noteCollectedAt','noteCollectedOk','function xhsCollectedHour(','function xhsCollectedBadge(','class="xhs-collection-badge ok"','<th>聚光</th><th>留资</th><th>笔记</th>','xhs-note-count-table','xhs-view-count-table','function xhsMetricTotal(account,weeks,field)','<th>汇总</th>','xhs-daily-reading-chart','id="xhs-leads"','xhs-goal-table','xhs-lead-opened-table','xhs-lead-copied-table','function xhsLeadWeekHeading(','xhs-week-day-badge','id="xhs-lead-details"','xhs-lead-detail-account','xhs-lead-detail-table','id="xhs-ad-flow"','xhs-ad-account-table','xhs-ad-note-table','id="xhs-ad-start-date"','id="xhs-ad-end-date"','function xhsAdPrepareDateControls(','id="xhs-ad-team-filter"','id="xhs-ad-account-filter"','id="xhs-ad-matrix-head"','function renderXhsAdChart(','function renderXhsAdFlow()','function renderXhsAccountStatus()','function xhsGoalCell(','function renderXhsLeads()','function renderXhsLeadDetails()','"xhsAccountAudit"','"targetMonth"','"targets"','"dailyRows"','"xhsLeads"','"xhsAdFlow"']
required=[{'id="xhs-ad-matrix-head"':'class="xhs-ad-matrix-head"'}.get(marker,marker) for marker in required]
required += ['data-dashboard-view="contract-details"','id="contract-details"','id="contract-detail-new-table"','id="contract-detail-renewal-table"','id="contract-detail-other-table"','id="contract-detail-checkout-table"','function renderMonthlyContractDetails()','"monthlyDetails"','"signSource"','"otherCount"','"otherRevenue"']
required += ['id="xhs-ad-spend-chart-wrap"','id="xhs-ad-spend-chart"','id="xhs-ad-spend-tooltip"','id="xhs-ad-leads-chart-wrap"','id="xhs-ad-leads-chart"','id="xhs-ad-leads-tooltip"','xhs-ad-spend-guide','xhs-ad-leads-guide','class="panel xhs-ad-detail-panel"','class="xhs-ad-date-tools xhs-ad-detail-date-tools"','id="xhs-ad-week-month-filter"','id="xhs-ad-week-team-filter"','id="xhs-ad-week-account-filter"','id="xhs-ad-week-owner-filter"','id="xhs-ad-week-summary"','id="xhs-ad-week-account-summary"','id="xhs-ad-week-owner-summary"','id="xhs-ad-week-head"','id="xhs-ad-week-table"','id="xhs-ad-owner-week-head"','id="xhs-ad-owner-week-table"','function xhsAdPrepareWeekControls(','function xhsAdWeekPeriods(','function xhsAdWeekDimension(','function renderXhsAdWeeklyTable(','function renderXhsAdDetailTables(','function xhsAdSummarizeAccountRows(','function xhsAdSummarizeOwnerRows(','function renderXhsAdSingleChart(','function bindXhsAdChartHover(','汇总为每个投流账号一行','汇总为每个归属账号一行']
required += ['id="xhs-traffic"','data-dashboard-view="xhs-traffic"','id="xhs-traffic-updated"','id="xhs-traffic-total"','id="xhs-traffic-paid"','id="xhs-traffic-organic"','id="xhs-traffic-organic-rate"','id="xhs-traffic-spend"','id="xhs-traffic-start-date"','id="xhs-traffic-end-date"','id="xhs-traffic-date-reset"','id="xhs-traffic-trend-chart"','id="xhs-traffic-trend-summary"','id="xhs-traffic-decline-grid"','id="xhs-traffic-table"','id="xhs-traffic-prev-page"','id="xhs-traffic-page-status"','id="xhs-traffic-next-page"','id="xhs-traffic-week-month"','id="xhs-traffic-week-table"','id="xhs-traffic-team-table"','class="panel-note xhs-traffic-footnote"','仅汇总两套数据共同覆盖日期','xhsTrafficState.initialized','function xhsTrafficBuildAccountRows(','adContent.ownerRows || []','function xhsTrafficBuildRows(','organicLeads=totalLeads-paidLeads','function xhsTrafficWeekPeriods(','function xhsTrafficDeclineAccounts(','function renderXhsTrafficTrendChart(','function renderXhsTrafficTrends(','function renderXhsTrafficWeekTable(','function renderXhsTrafficTeamTable(','function renderXhsTraffic(']
required += ['最近7个有数据日期对比此前7个有数据日期','至少需要14个有数据日期','上7日均','近7日均','日均变化','日均上升','日均持平','fourteen.slice(0,7)','fourteen.slice(7,14)','function xhsTrafficDeclineAccounts(accountRows,metric=','Number(row[metric] || 0)',').filter(Boolean).sort((a,b) => a.delta-b.delta']
required += ['id="xhs-note-published"','data-dashboard-view="xhs-note-published"','id="xhs-note-published-account"','id="xhs-note-published-count"','id="xhs-note-published-table"','function renderXhsNotePublished()','"xhsNotePublished"']
required += ['id="xhs-note-published-type"','图文数量','视频数量','待识别数量','graphicCount','videoCount','pendingCount']
required += ['数据明细','笔记发布明细','留资数据明细','聚光投放明细','data-mobile-menu="xhs-details"','data-mobile-submenu="xhs-details"','id="xhs-ad-details"','data-dashboard-view="xhs-ad-details"','id="xhs-ad-details-updated"','id="xhs-ad-detail-account-filter"','id="xhs-ad-detail-start-date"','id="xhs-ad-detail-end-date"','id="xhs-ad-detail-date-reset"','id="xhs-ad-details-count"','id="xhs-ad-details-table"','function renderXhsAdDetails(']
required += ['class="desktop-home-link"','class="desktop-nav-groups"','class="desktop-nav-group"','class="desktop-module-toggle"','aria-expanded="false"','aria-expanded="true"']
required += ['<div class="label">本月退房</div>','id="contract-checkout-definition">退租/（实际退/续租）','function checkoutDisplay(','checkoutActualDepartureCount']
required += ['data-desktop-module="customer"','data-desktop-menu="customer"','data-mobile-module="customer"','data-mobile-menu="customer"','id="customer-data"','data-dashboard-view="customer-data"','id="customer-data-updated"','id="customer-daily-table"','id="customer-funnel-table"','function renderCustomerData()','"customerData"']
required += ['data-desktop-module="meters"','data-desktop-menu="meters"','data-mobile-module="meters"','data-mobile-menu="meters"','id="meter-management"','data-dashboard-view="meter-management"','id="meter-collection-status"','id="meter-keep-table"','id="meter-negative-table"','id="meter-offline-table"','function renderMeterManagement()','"meterManagement"','data-label="保电状态"','keepText=row.keepElectric?\'已保电\':\'未保电\'']
required += ['function xhsNoteCountClass(','function xhsMetricCell(','xhs-note-count-green','xhs-note-count-yellow','xhs-note-count-pink','xhs-note-count-red','if(count>=6)','if(count===5)','if(count===4)']
required += ['data-dashboard-view="overview-new"','id="overview-new"','function renderOverviewNew()','"overviewNew"','overview-new-short-rent','overview-new-rate-comprehensive','overview-new-validation']
required += ['id="overview-contract-activity"','overview-contract-today-new-sign','overview-contract-yesterday-reservation','overview-contract-week-actual-checkout','overview-contract-month-new-sign','overview-contract-month-reservation','overview-contract-month-renewal','overview-contract-month-actual-checkout','.overview-contract-kpis .hint{display:none}','"contractActivity"','function overviewContractRangeLabel(','function overviewContractMetricDisplay(','newSignRevenue','reservationRevenue','renewalRevenue','monthCoreMatched','detailCountsMatched','id="overview-contract-detail-modal"','function openOverviewContractDetails(','overview-contract-drill-card']
required += ['id="business-trend"','data-dashboard-view="business-trend"','id="business-trend-chart"','id="business-trend-summary"','function businessTrendDateLabel(','function renderBusinessTrend()','"businessTrend"','newSignCount','reservationCount','最新日期在左','堆叠面积图','新签数量（底层）','预定数量（上层）','business-trend-area new-sign','business-trend-area reservation','const totalAt=','areaPath(totalAt','business-trend-line total','business-trend-value total','business-trend-value new-sign','newSign!==total','business-trend-grid vertical','month-boundary','labelY=Math.max(108,pointY-9)','business-trend-month-band','business-trend-week-band','weekRangeCoverage','monthRangeCoverage']
required += ['id="checkout-reason-trend-card"','id="checkout-reason-trend-summary"','id="checkout-reason-trend-chart"','function renderCheckoutReasonTrend(','reasonRows','reasonCategories','displayedReasonCategories','reasonRanges','reasonMonths','reasonField','expiryCount','breachCount','renewalCount','otherCount','displayTotalCount','实际退租原因趋势','到期（底层）','续租（中层）','违约（上层）','总数折线','checkout-reason-area expiry','checkout-reason-area renewal','checkout-reason-area breach','checkout-reason-total-line','checkout-reason-month-band','checkout-reason-week-rate','续租率']
required += ['id="contract-daily-trend-panel"','class="panel contract-daily-trend-panel"','id="contract-daily-chart-wrap"','id="contract-daily-chart"','function renderDailyLineChart()','renderBusinessTrend(); renderDailyLineChart(); renderCheckoutTrends();']
required += ['id="checkout-trend-grid"','id="checkout-trend-future-chart"','id="checkout-trend-past-chart"','id="checkout-trend-future-summary"','id="checkout-trend-past-summary"','function renderCheckoutTrends()','function renderCheckoutTrendChart(','"checkoutTrends"','"ranges"','"periodKey"','"week"','"label"','futureNearestFirst','pastNewestFirst','pastRangeCoverage','pastRangeTotalsMatched','futureRangeCoverage','futureRangeTotalsMatched','未来30天','过去30天','checkoutLabelY=Math.max(18,pointY-9)','checkout-trend-range-band','checkout-trend-range-total','compactLabel=compactMonth','天合计']
required += ['function weekKeyFromLabel(','function weekDayBadgeInfo(','function weekHeading(','id="project-checkout-head"','id="building-checkout-head"']
required += ['id="xhs-reading-decline-grid"','function renderXhsReadingDeclines()','function xhsDeclineSparkline(','"accountDailyReading"','上7个有效采集日平均－近7个有效采集日平均']
required += ['id="xhs-trends"','data-dashboard-view="xhs-trends"','id="xhs-trends-updated"','id="xhs-trend-title"','id="xhs-trend-metric"','id="xhs-trend-main-content"','id="xhs-trend-decline-content"','id="xhs-traffic-decline-summary"','value="totalLeads"','value="organicLeads"','value="paidLeads"','const labelStep=1','xhsTrafficDeclineAccounts(accountRows.filter((row) => row.date!==today),metric)','renderXhsTrafficTrendChart(\'xhs-traffic-decline-chart-\'+index,account.rows,true,metric)','function renderXhsTrafficTrendChart(svgId,rows,compact=false,metric=','已隐藏当日未完整数据','每个日期与数值均完整展示','rows.filter((row) => row.date!==today).slice(0,30)','Number(row.date.slice(5,7))+\'-\'+Number(row.date.slice(8,10))']
if any(x not in rendered for x in required): raise RuntimeError("Full dashboard style validation failed")
if rendered.count('data-dashboard-view="xhs-traffic"') < 2: raise RuntimeError("XHS traffic menu must exist on desktop and mobile")
if rendered.count('data-dashboard-view="xhs-trends"') < 2: raise RuntimeError("XHS trends menu must exist on desktop and mobile")
trends_section=re.search(r'<section class="section" id="xhs-trends".*?</section>\s*<section class="section" id="xhs-traffic"',rendered,re.S)
traffic_section=re.search(r'<section class="section" id="xhs-traffic".*?</section>\s*<section class="section" id="xhs-ad-flow"',rendered,re.S)
if not trends_section or 'xhs-traffic-trend-chart' not in trends_section.group(0) or 'xhs-traffic-decline-grid' not in trends_section.group(0) or trends_section.group(0).count('<article class="panel">') != 1: raise RuntimeError("XHS trend content must share one panel")
if 'value="declineAccounts"' in trends_section.group(0) or 'hidden' in re.search(r'<div class="xhs-trend-decline-content"[^>]*>',trends_section.group(0)).group(0): raise RuntimeError("XHS decline accounts must stay visible and follow the selected metric")
if not traffic_section or 'xhs-traffic-trend-chart' in traffic_section.group(0) or 'xhs-traffic-decline-grid' in traffic_section.group(0): raise RuntimeError("XHS trend panels must be removed from traffic view")
if rendered.count('data-dashboard-view="xhs-note-published"') < 2: raise RuntimeError("XHS note-published menu must exist on desktop and mobile")
if rendered.count('data-dashboard-view="xhs-lead-details"') < 2: raise RuntimeError("XHS lead-details menu must exist on desktop and mobile")
if rendered.count('data-dashboard-view="xhs-ad-details"') < 2: raise RuntimeError("XHS ad-details menu must exist on desktop and mobile")
if rendered.count('data-dashboard-view="customer-data"') < 2: raise RuntimeError("Customer data menu must exist on desktop and mobile")
if rendered.count('data-dashboard-view="meter-management"') < 2: raise RuntimeError("Meter-management menu must exist on desktop and mobile")
if rendered.count('data-dashboard-view="overview-new"') < 2: raise RuntimeError("Overview-new menu must exist on desktop and mobile")
if rendered.count('data-dashboard-view="business-trend"') < 2: raise RuntimeError("Business-trend menu must exist on desktop and mobile")
customer_rows=payload["customerData"].get("dailyRows",[])
if customer_rows:
    customer_fields=("published","reading","inbound","leads","wechatAdds","actualTours","signed","deposits")
    if len(customer_rows)!=7 or len({row.get("date") for row in customer_rows})!=7 or customer_rows!=sorted(customer_rows,key=lambda row:row["date"]): raise RuntimeError("Customer data seven-day coverage invalid")
    if any(payload["customerData"]["totals"].get(field)!=sum(row[field] for row in customer_rows) for field in customer_fields): raise RuntimeError("Customer data totals do not reconcile")
account_audit_rows=payload["xhsAccountAudit"].get("accounts",[])
if len(account_audit_rows)!=len(XHS_ACCOUNTS) or any(not all(key in row for key in ("adCollectedAt","adCollectedOk","leadCollectedAt","leadCollectedOk","noteCollectedAt","noteCollectedOk")) for row in account_audit_rows): raise RuntimeError("XHS account collection timestamps invalid")
note_published_rows=payload["xhsNotePublished"].get("rows",[])
if len(note_published_rows)<232 or len({(row.get("profile"),row.get("publishedDate")) for row in note_published_rows})!=len(note_published_rows): raise RuntimeError("XHS note-published history coverage invalid")
if any(int(row.get("graphicCount") or 0)+int(row.get("videoCount") or 0)+int(row.get("pendingCount") or 0)!=int(row.get("publishedCount") or 0) for row in note_published_rows): raise RuntimeError("XHS note type totals do not reconcile")
ad_flow=payload["xhsAdFlow"]
ad_detail_rows=ad_flow.get("accountRows",[])
if not ad_detail_rows or any(not all(key in row for key in ("date","accountName","spend","opened","leads")) for row in ad_detail_rows): raise RuntimeError("XHS ad-detail history coverage invalid")
if ad_flow.get("historySource")=="immutable-history/ad-note-daily.csv":
    if ad_flow.get("date")!=ad_flow.get("historyMaxDate") or ad_flow.get("date")!=max(row["date"] for row in ad_detail_rows): raise RuntimeError("XHS ad immutable freshness validation failed")
    start_text,end_text=ad_flow["periodLabel"].split(" 至 ",1)
    expected_days=(datetime.strptime(end_text,"%Y-%m-%d").date()-datetime.strptime(start_text,"%Y-%m-%d").date()).days+1
    if len(ad_detail_rows)!=expected_days*len(XHS_ACCOUNTS): raise RuntimeError("XHS ad immutable account-day coverage invalid")
if 'data-dashboard-view="checkout"' in rendered: raise RuntimeError("Legacy checkout navigation detected")
mobile_yuxiaor=re.search(r'<nav class="mobile-menu mobile-secondary-nav" data-mobile-menu="yuxiaor".*?</nav>',rendered,re.S)
if mobile_yuxiaor and 'occupancy-ziyin' in mobile_yuxiaor.group(0): raise RuntimeError("Ziyin occupancy menu must stay desktop-only")
if 'id="xhs-lead-inbound-table"' in rendered: raise RuntimeError("Private-message inbound table must stay hidden")
if len(building_rows)<50 or project_data[0]["rooms"]!=692: raise RuntimeError("Data reconciliation failed")
if project_data[0]["lockCount"]!=lock_total: raise RuntimeError("Lock count reconciliation failed")
if project_data[0]["rentableVacancyCount"]+project_data[0]["unrentableVacancyCount"]!=project_data[0]["vacancyCount"]: raise RuntimeError("Vacancy availability reconciliation failed")
if project_data[0]["comprehensiveCount"]!=project_data[0]["occupiedCount"]+project_data[0]["preorderCount"]+project_data[0]["moveInCount"]: raise RuntimeError("Comprehensive occupancy reconciliation failed")
if house_monitor["occupied"]+house_monitor["rentable"]+house_monitor["unrentable"]+house_monitor["unknownStatus"]!=project_data[0]["rooms"]: raise RuntimeError("Room status reconciliation failed")
if house_monitor["preordered"]+house_monitor["movingIn"]>house_monitor["unrentable"]: raise RuntimeError("Unavailable room classification failed")
output_path.parent.mkdir(parents=True,exist_ok=True)
tmp=output_path.with_suffix(".tmp"); tmp.write_text(rendered,encoding="utf-8"); tmp.replace(output_path)
print(json.dumps({"dataDate":payload["dataDate"],"rooms":project_data[0]["rooms"],"lockCount":lock_total,"lockField":"锁房备注","rentableVacancyCount":house_monitor["rentable"],"unrentableVacancyCount":house_monitor["unrentable"],"mappedUnknownCount":house_monitor["mappedUnknown"],"occupiedCount":house_monitor["occupied"],"preorderCount":house_monitor["preordered"],"moveInCount":house_monitor["movingIn"],"unknownStatusCount":house_monitor["unknownStatus"],"comprehensiveDefinition":"在租中/已出租+空房不可租中的已预订和将搬入","vacancyField":"状态（未知状态映射为其他锁房）","buildings":len(building_rows),"projects":len(project_data),"style":"latest-full"},ensure_ascii=False))
