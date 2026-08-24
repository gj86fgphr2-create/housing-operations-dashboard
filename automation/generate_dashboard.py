from __future__ import annotations

import json
import os
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(os.environ.get('YUXIAOR_ROOT', '/opt/yuxiaor-automation'))
DATA = ROOT / 'data' / 'current'
SITE = ROOT / 'site'


def atomic_write_text(target: Path, content: str) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f'.{target.name}.{os.getpid()}.tmp')
    try:
        temporary.write_text(content, encoding='utf-8')
        os.chmod(temporary, 0o644)
        os.replace(temporary, target)
    finally:
        temporary.unlink(missing_ok=True)

def rows(file_name: str, header_row: int = 3):
    ws = load_workbook(DATA / file_name, read_only=True, data_only=True).worksheets[0]
    values = ws.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(values, None)
    headers = [str(v or '').strip() for v in next(values, [])]
    for row in values:
        item = {headers[i]: value for i, value in enumerate(row) if i < len(headers) and headers[i]}
        if any(value not in (None, '') for value in item.values()):
            yield item

def text(value):
    return str(value or '').strip()

def number(value):
    try:
        return float(str(value or 0).replace(',', '').replace('¥', ''))
    except ValueError:
        return 0.0

def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = text(value).split(' ')[0].replace('/', '-')
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return None

today = datetime.now().astimezone().date()
month_start = today.replace(day=1)
previous_end = month_start - timedelta(days=1)
previous_start = previous_end.replace(day=1)

houses = list(rows('房源详情.xlsx'))
renting = list(rows('在租中合同.xlsx'))
moving = list(rows('将搬入合同.xlsx'))
retired = list(rows('已退租合同.xlsx'))
contracts = renting + moving + retired

def house_key(row):
    return text(row.get('房源ID')) or text(row.get('房间ID')) or text(row.get('房源流水号'))

def building_name(row):
    return text(row.get('小区/公寓')) or text(row.get('房源地址')).split('座')[0] + '座'

house_by_key = {house_key(row): row for row in houses if house_key(row)}
renting_by_key = {house_key(row): row for row in renting if house_key(row)}
moving_by_key = {house_key(row): row for row in moving if house_key(row)}

building_projects = defaultdict(Counter)
for row in renting + moving:
    building_projects[building_name(row)][text(row.get('所属部门')) or '未分配项目'] += 1

def project_name(building):
    counts = building_projects.get(building)
    return counts.most_common(1)[0][0] if counts else '未分配项目'

def checkout_stats(keys, room_count):
    dates = []
    for key in keys:
        row = renting_by_key.get(key)
        if row:
            d = as_date(row.get('退租时间'))
            if d:
                dates.append(d)
    next7 = sum(today <= d <= today + timedelta(days=6) for d in dates)
    days8to14 = sum(today + timedelta(days=7) <= d <= today + timedelta(days=13) for d in dates)
    week_end = today + timedelta(days=6 - today.weekday())
    week = sum(today <= d <= week_end for d in dates)
    month = sum(d.year == today.year and d.month == today.month and d >= today for d in dates)
    denom = room_count or 1
    return {
        'next7': {'count': next7, 'rate': next7 / denom},
        'days8to14': {'count': days8to14, 'rate': days8to14 / denom},
        'week': {'count': week, 'rate': week / denom},
        'month': {'count': month, 'rate': month / denom},
    }

def aggregate(name, group):
    keys = [house_key(row) for row in group if house_key(row)]
    rooms = len(group)
    occupied = sum(key in renting_by_key for key in keys)
    future = sum(key in moving_by_key for key in keys)
    vacancy = max(0, rooms - occupied)
    comprehensive = min(rooms, occupied + future)
    return {
        'name': name, 'rooms': rooms,
        'comprehensiveCount': comprehensive, 'comprehensiveRate': comprehensive / rooms if rooms else 0,
        'occupiedCount': occupied, 'occupancyRate': occupied / rooms if rooms else 0,
        'vacancyCount': vacancy, 'vacancyRate': vacancy / rooms if rooms else 0,
        'lockCount': 0, 'lockRate': 0, 'preorderCount': 0,
        'checkout': checkout_stats(keys, rooms),
    }

by_building = defaultdict(list)
by_project = defaultdict(list)
for house in houses:
    building = building_name(house)
    by_building[building].append(house)
    by_project[project_name(building)].append(house)

building_data = [aggregate(name, group) for name, group in sorted(by_building.items())]
project_data = [aggregate('全部房源汇总', houses)] + [aggregate(name, group) for name, group in sorted(by_project.items())]

dedup = {}
for row in contracts:
    dedup[text(row.get('合同编号')) or f'row-{len(dedup)}'] = row
contract_rows = list(dedup.values())

def signing_category(row):
    source = text(row.get('签约来源'))
    if source == '新签':
        return 'new'
    if source == '续租':
        return 'renewal'
    if source in {'换房', '重签'}:
        return 'other'
    return 'uncategorized'

def signed_in(row, start, end):
    d = as_date(row.get('签约时间'))
    return bool(d and start <= d <= end)

def actual_checkout_in(row, start, end):
    d = as_date(row.get('预退/实退'))
    return bool(d and start <= d <= end and text(row.get('退租原因')) not in {'换房清算'})

def period_summary(label, start, end):
    signed = [r for r in contract_rows if signed_in(r, start, end)]
    new_rows = [r for r in signed if signing_category(r) == 'new']
    renew_rows = [r for r in signed if signing_category(r) == 'renewal']
    other_rows = [r for r in signed if signing_category(r) == 'other']
    return {
        'period': label, 'startDate': start.isoformat(), 'endDate': end.isoformat(),
        'newCount': len(new_rows), 'newRevenue': sum(number(r.get('总租金')) for r in new_rows),
        'renewalCount': len(renew_rows), 'renewalRevenue': sum(number(r.get('总租金')) for r in renew_rows),
        'otherCount': len(other_rows), 'otherRevenue': sum(number(r.get('总租金')) for r in other_rows),
    }

current = period_summary('本月', month_start, today)
previous = period_summary('上月', previous_start, previous_end)
week_periods = []
month_end = month_start.replace(
    month=month_start.month % 12 + 1,
    year=month_start.year + (1 if month_start.month == 12 else 0),
) - timedelta(days=1)
for index, start_day in enumerate((1, 8, 15, 22, 29), start=1):
    start_day_date = month_start.replace(day=start_day)
    end_day = month_start.replace(day=min(start_day + 6, month_end.day))
    label = f'{today.month}-WEEK-{index}' if index < 5 else f'{today.month}-WEEK-END'
    week_periods.append(period_summary(label, start_day_date, end_day))

people = defaultdict(lambda: {'newCount': 0, 'newRevenue': 0, 'renewalCount': 0, 'renewalRevenue': 0, 'otherCount': 0, 'otherRevenue': 0})
for row in [r for r in contract_rows if signed_in(r, month_start, today)]:
    name = text(row.get('签约人')) or '未填写'
    key = signing_category(row)
    if key == 'uncategorized':
        continue
    people[name][f'{key}Count'] += 1
    people[name][f'{key}Revenue'] += number(row.get('总租金'))

daily = []
for offset in range(45, -1, -1):
    d = today - timedelta(days=offset)
    signed = [r for r in contract_rows if as_date(r.get('签约时间')) == d]
    new_count = sum(signing_category(r) == 'new' for r in signed)
    renewal_count = sum(signing_category(r) == 'renewal' for r in signed)
    other_count = sum(signing_category(r) == 'other' for r in signed)
    checkout = sum(actual_checkout_in(r, d, d) for r in retired)
    daily.append({'date': d.isoformat(), 'newSign': new_count, 'renewal': renewal_count, 'other': other_count, 'checkout': checkout, 'total': new_count - checkout})

project_monthly = defaultdict(lambda: {'newCount': 0, 'renewalCount': 0, 'otherCount': 0, 'actualCheckoutCount': 0, 'checkoutRenewalCount': 0})
for row in contract_rows:
    project = text(row.get('所属部门')) or '未分配项目'
    if signed_in(row, month_start, today):
        key = signing_category(row)
        if key != 'uncategorized':
            project_monthly[project][f'{key}Count'] += 1
for row in retired:
    if actual_checkout_in(row, month_start, today):
        project = text(row.get('所属部门')) or '未分配项目'
        project_monthly[project]['actualCheckoutCount'] += 1
        if text(row.get('退租原因')) == '续租':
            project_monthly[project]['checkoutRenewalCount'] += 1

checkout_current = [r for r in retired if actual_checkout_in(r, month_start, today)]
checkout_previous = [r for r in retired if actual_checkout_in(r, previous_start, previous_end)]

def checkout_breakdown(rows):
    renewal_count = sum(text(row.get('退租原因')) == '续租' for row in rows)
    return {
        'actualCheckoutCount': len(rows),
        'checkoutActualDepartureCount': len(rows) - renewal_count,
        'checkoutRenewalCount': renewal_count,
    }

targets = [
    {'period': f'{today.month}-WEEK-1', 'label': 'WEEK-1', 'newCount': 27, 'renewalCount': 14},
    {'period': f'{today.month}-WEEK-2', 'label': 'WEEK-2', 'newCount': 32, 'renewalCount': 17},
    {'period': f'{today.month}-WEEK-3', 'label': 'WEEK-3', 'newCount': 39, 'renewalCount': 21},
    {'period': f'{today.month}-WEEK-4', 'label': 'WEEK-4', 'newCount': 43, 'renewalCount': 23},
    {'period': f'{today.month}-WEEK-END', 'label': 'WEEK-END', 'newCount': 9, 'renewalCount': 5},
]

payload = {
    'dataDate': today.isoformat(), 'generatedDate': today.isoformat(),
    'projectData': project_data, 'buildingData': building_data,
    'baseProjectNames': [row['name'] for row in project_data],
    'contractStats': {
        'currentMonth': {**current, **checkout_breakdown(checkout_current)},
        'previousMonth': {**previous, **checkout_breakdown(checkout_previous)},
        'periods': [current, previous] + week_periods, 'targets': targets,
        'people': [{'name': name, **values} for name, values in people.items()],
        'projectMonthly': [{'name': name, **values} for name, values in project_monthly.items()],
        'daily': daily,
    },
}

SITE.mkdir(parents=True, exist_ok=True)
template = (ROOT / 'app' / 'dashboard-template.html').read_text(encoding='utf-8')
output = template.replace('__EMBEDDED_DATA__', json.dumps(payload, ensure_ascii=False, separators=(',', ':')))
atomic_write_text(SITE / 'index.html', output)
atomic_write_text(SITE / 'dashboard-data.json', json.dumps(payload, ensure_ascii=False, indent=2))
print(json.dumps({'rooms': len(houses), 'renting': len(renting), 'moving': len(moving), 'retired': len(retired), 'buildings': len(building_data), 'site': str(SITE / 'index.html')}, ensure_ascii=False))
