#!/usr/bin/env python3
"""Build validated Xiaohongshu Aurora history snapshots and Excel exports."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ACCOUNTS = [
    {"profile": "account-02", "name": "广州大学城租房-研寓", "team": "管家团队"},
    {"profile": "account-03", "name": "广州研寓租房大学城", "team": "管家团队"},
    {"profile": "account-04", "name": "大学城捞房长短租随意", "team": "运营团队"},
    {"profile": "account-05", "name": "暴走大学城探房版", "team": "运营团队"},
    {"profile": "account-06", "name": "广州大学城-研舍公寓", "team": "管家团队"},
    {"profile": "account-07", "name": "广州大学城租房-维特", "team": "管家团队"},
    {"profile": "account-08", "name": "大学城租房 | 研舍", "team": "管家团队"},
    {"profile": "account-09", "name": "番禺大学城租房-尚维特", "team": "管家团队"},
]
ACCOUNT_BY_PROFILE = {row["profile"]: row for row in ACCOUNTS}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--raw-root", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    parser.add_argument("--start-date", required=True)
    parser.add_argument("--end-date", required=True)
    parser.add_argument("--allow-incomplete", action="store_true")
    return parser.parse_args()


def date_range(start: str, end: str) -> list[str]:
    first = dt.date.fromisoformat(start)
    last = dt.date.fromisoformat(end)
    if first > last:
        raise ValueError("start date is after end date")
    return [(first + dt.timedelta(days=offset)).isoformat() for offset in range((last - first).days + 1)]


def atomic_write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary = tempfile.mkstemp(prefix=f".{path.name}.", dir=path.parent)
    try:
        with os.fdopen(handle, "w", encoding="utf-8", newline="") as stream:
            stream.write(content)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)


def money(value: Any) -> float:
    try:
        return round(float(value or 0), 2)
    except (TypeError, ValueError):
        return 0.0


def integer(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def owner_team(name: str) -> str:
    return "运营团队" if "大学城捞房" in name or "暴走大学城" in name else "管家团队"


def load_sources(raw_root: Path) -> list[tuple[Path, dict[str, Any]]]:
    sources = []
    corrected_sources = sorted(raw_root.glob("account-*/corrected-range/*.json"))
    candidates = corrected_sources if len(corrected_sources) == len(ACCOUNTS) else sorted(raw_root.rglob("*.json"))
    for path in candidates:
        if "probe" in path.parts:
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if isinstance(payload, dict) and isinstance(payload.get("accounts"), list):
            sources.append((path, payload))
    return sources


def build_history(raw_root: Path, start: str, end: str) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    dates = date_range(start, end)
    expected_profiles = set(ACCOUNT_BY_PROFILE)
    account_periods: dict[str, list[tuple[str, str, dict[str, Any], str]]] = defaultdict(list)
    facts: dict[tuple[str, str, str], dict[str, Any]] = {}

    for path, payload in load_sources(raw_root):
        payload_start = str(payload.get("start_date") or payload.get("date") or "")
        payload_end = str(payload.get("end_date") or payload.get("date") or "")
        if not payload_start or not payload_end or payload_end < start or payload_start > end:
            continue
        generated = str(payload.get("generated_at") or path.stat().st_mtime_ns)
        for account in payload.get("accounts", []):
            profile = str(account.get("profile") or "")
            if profile not in expected_profiles:
                continue
            account_periods[profile].append((payload_start, payload_end, account, generated))
            for row in account.get("rows") or []:
                row_date = str(row.get("date") or payload.get("date") or "")
                note_id = str(row.get("note_id") or "")
                if not (start <= row_date <= end and note_id):
                    continue
                key = (row_date, profile, note_id)
                fact = {
                    "date": row_date,
                    "profile": profile,
                    "account_name": str(account.get("account_name") or ACCOUNT_BY_PROFILE[profile]["name"]),
                    "team": ACCOUNT_BY_PROFILE[profile]["team"],
                    "note_id": note_id,
                    "spend": money(row.get("spend")),
                    "private_message_opens": integer(row.get("private_message_opens")),
                    "private_message_leads": integer(row.get("private_message_leads")),
                    "owner_account_name": str(row.get("owner_account_name") or "").strip(),
                    "owner_user_id": str(row.get("owner_user_id") or "").strip(),
                    "owner_source": str(row.get("owner_source") or ""),
                    "owner_status": str(row.get("owner_status") or "unresolved"),
                    "collected_at": generated,
                    "source_file": str(path),
                }
                existing = facts.get(key)
                if existing is None or generated >= existing["collected_at"]:
                    facts[key] = fact

    coverage = []
    account_daily = []
    facts_by_profile_date: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for fact in facts.values():
        facts_by_profile_date[(fact["profile"], fact["date"])].append(fact)

    for account_def in ACCOUNTS:
        profile = account_def["profile"]
        for day in dates:
            rows = facts_by_profile_date.get((profile, day), [])
            covering_periods = [period for period in account_periods.get(profile, []) if period[0] <= day <= period[1]]
            period = max(covering_periods, key=lambda item: item[3]) if covering_periods else None
            covered = period is not None
            source_status = str(period[2].get("status") or "missing") if covered else "missing"
            if source_status == "ok":
                quality_status = "有数据" if rows else "已确认零消耗"
            elif source_status == "not_logged_in":
                quality_status = "未登录"
            else:
                quality_status = "失败" if covered else "缺失"
            error = str(period[2].get("error") or "") if covered else "无覆盖采集结果"
            coverage.append({
                "date": day,
                "profile": profile,
                "accountName": account_def["name"],
                "team": account_def["team"],
                "status": quality_status,
                "sourceStatus": source_status,
                "rowCount": len(rows),
                "error": error,
            })
            spend = round(sum(row["spend"] for row in rows), 2)
            opened = sum(row["private_message_opens"] for row in rows)
            leads = sum(row["private_message_leads"] for row in rows)
            account_daily.append({
                "date": day,
                "profile": profile,
                "accountName": account_def["name"],
                "team": account_def["team"],
                "noteCount": len(rows),
                "spend": spend,
                "opened": opened,
                "averageOpenCost": round(spend / opened, 2) if opened else None,
                "leads": leads,
                "averageLeadCost": round(spend / leads, 2) if leads else None,
                "status": "ok" if source_status == "ok" else source_status,
                "statusLabel": quality_status,
                "error": error,
            })

    owner_groups: dict[tuple[str, str], dict[str, Any]] = {}
    for fact in facts.values():
        owner_name = fact["owner_account_name"]
        owner_id = fact["owner_user_id"]
        key = (fact["date"], owner_id or owner_name or "unresolved")
        group = owner_groups.setdefault(key, {
            "date": fact["date"],
            "ownerAccountName": owner_name,
            "ownerUserId": owner_id,
            "team": owner_team(owner_name),
            "noteCount": 0,
            "spend": 0.0,
            "opened": 0,
            "leads": 0,
            "ownerStatus": "confirmed",
            "ownerStatusLabel": "已确认",
        })
        group["noteCount"] += 1
        group["spend"] += fact["spend"]
        group["opened"] += fact["private_message_opens"]
        group["leads"] += fact["private_message_leads"]
        if not owner_name or not owner_id or fact["owner_status"] != "confirmed":
            group["ownerStatus"] = "unresolved"
            group["ownerStatusLabel"] = "待确认"

    owner_daily = list(owner_groups.values())
    for row in owner_daily:
        row["spend"] = round(row["spend"], 2)
        row["averageOpenCost"] = round(row["spend"] / row["opened"], 2) if row["opened"] else None
        row["averageLeadCost"] = round(row["spend"] / row["leads"], 2) if row["leads"] else None

    account_daily.sort(key=lambda row: (row["date"], row["spend"], row["profile"]), reverse=True)
    owner_daily.sort(key=lambda row: (row["date"], row["spend"], row["ownerAccountName"]), reverse=True)
    fact_rows = sorted(facts.values(), key=lambda row: (row["date"], row["profile"], row["note_id"]))
    unresolved = [row for row in fact_rows if not row["owner_account_name"] or not row["owner_user_id"] or row["owner_status"] != "confirmed"]
    incomplete = [row for row in coverage if row["status"] not in {"有数据", "已确认零消耗"}]
    snapshot = {
        "schema": "xhs-ad-history-v1",
        "generated_at": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "start_date": start,
        "end_date": end,
        "date": end,
        "periodLabel": f"{start} 至 {end}",
        "aggregation": "DAY",
        "accountRows": account_daily,
        "ownerRows": owner_daily,
        "coverage": coverage,
        "coverageExpected": len(ACCOUNTS) * len(dates),
        "coverageComplete": len(incomplete) == 0,
        "incompleteCount": len(incomplete),
        "totalNoteCount": len(fact_rows),
        "totalSpend": round(sum(row["spend"] for row in fact_rows), 2),
        "totalOpened": sum(row["private_message_opens"] for row in fact_rows),
        "totalLeads": sum(row["private_message_leads"] for row in fact_rows),
        "unresolvedOwnerCount": len(unresolved),
    }
    return snapshot, fact_rows


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    buffer = []
    import io
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    atomic_write_text(path, "\ufeff" + stream.getvalue())


def style_sheet(sheet, widths: list[int]) -> None:
    header_fill = PatternFill("solid", fgColor="DCEAF4")
    header_font = Font(bold=True, color="102B4E")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def write_excel(path: Path, snapshot: dict[str, Any]) -> None:
    workbook = Workbook()
    account_sheet = workbook.active
    account_sheet.title = "投流账号消耗"
    headers = ["日期", "团队", "投流账号", "笔记数", "消耗", "私信开口数", "平均开口成本", "私信留资数", "平均留资成本", "采集状态"]
    account_sheet.append(headers)
    for row in snapshot["accountRows"]:
        if float(row["spend"] or 0) <= 0:
            continue
        account_sheet.append([row["date"], row["team"], row["accountName"], row["noteCount"], row["spend"], row["opened"], row["averageOpenCost"], row["leads"], row["averageLeadCost"], row["statusLabel"]])
    style_sheet(account_sheet, [13, 12, 30, 10, 12, 14, 16, 14, 16, 14])

    owner_sheet = workbook.create_sheet("笔记归属账户投流汇总")
    owner_sheet.append(["日期", "团队", "归属账号", "笔记数", "消耗", "私信开口数", "平均开口成本", "私信留资数", "平均留资成本", "归属状态"])
    for row in snapshot["ownerRows"]:
        if float(row["spend"] or 0) <= 0:
            continue
        owner_sheet.append([row["date"], row["team"], row["ownerAccountName"], row["noteCount"], row["spend"], row["opened"], row["averageOpenCost"], row["leads"], row["averageLeadCost"], row["ownerStatusLabel"]])
    style_sheet(owner_sheet, [13, 12, 30, 10, 12, 14, 16, 14, 16, 14])

    audit_sheet = workbook.create_sheet("采集完整性报告")
    audit_sheet.append(["日期", "账号", "团队", "状态", "笔记行数", "错误"])
    for row in snapshot["coverage"]:
        audit_sheet.append([row["date"], row["accountName"], row["team"], row["status"], row["rowCount"], row["error"]])
    style_sheet(audit_sheet, [13, 30, 12, 16, 12, 60])

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    workbook.save(temporary)
    os.replace(temporary, path)


def main() -> int:
    args = parse_args()
    snapshot, facts = build_history(args.raw_root, args.start_date, args.end_date)
    args.output_root.mkdir(parents=True, exist_ok=True)
    atomic_write_text(args.output_root / "latest.json", json.dumps(snapshot, ensure_ascii=False, indent=2) + "\n")
    write_csv(args.output_root / "ad-note-facts.csv", facts, [
        "date", "profile", "account_name", "team", "note_id", "spend",
        "private_message_opens", "private_message_leads", "owner_account_name",
        "owner_user_id", "owner_source", "owner_status", "collected_at", "source_file",
    ])
    write_csv(args.output_root / "account-daily.csv", snapshot["accountRows"], list(snapshot["accountRows"][0]) if snapshot["accountRows"] else [])
    write_csv(args.output_root / "owner-daily.csv", snapshot["ownerRows"], list(snapshot["ownerRows"][0]) if snapshot["ownerRows"] else [])
    export = args.output_root / "exports" / f"聚光投流历史数据_{args.start_date}_至_{args.end_date}.xlsx"
    write_excel(export, snapshot)
    print(json.dumps({
        "snapshot": str(args.output_root / "latest.json"),
        "excel": str(export),
        "coverageExpected": snapshot["coverageExpected"],
        "incompleteCount": snapshot["incompleteCount"],
        "notes": snapshot["totalNoteCount"],
        "spend": snapshot["totalSpend"],
        "opened": snapshot["totalOpened"],
        "leads": snapshot["totalLeads"],
        "unresolved": snapshot["unresolvedOwnerCount"],
    }, ensure_ascii=False))
    if not args.allow_incomplete and (snapshot["incompleteCount"] or snapshot["unresolvedOwnerCount"]):
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
