#!/usr/bin/env python3
import json
import os
import re
import sys
import uuid
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


FILES = ("房源详情.xlsx", "在租中合同.xlsx", "将搬入合同.xlsx", "已退租合同.xlsx", "预定合同.xlsx")
DASHBOARD_URL = "https://gj86fgphr2-create.github.io/housing-operations-dashboard/#operations-brief"
EXPECTED_STATUS = {
    "在租中合同.xlsx": "在租中",
    "将搬入合同.xlsx": "将搬入",
    "已退租合同.xlsx": "已退租",
}


def text(value):
    return "" if value is None else str(value).strip()


def norm(value):
    return text(value).replace(" ", "")


def iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    value = text(value)
    match = re.match(r"(\d{4})[-/.年](\d{1,2})[-/.月](\d{1,2})", value)
    return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}" if match else ""


def column(headers, name):
    cleaned = [norm(value) for value in headers]
    return cleaned.index(norm(name))


def money(value):
    raw = text(value).replace(",", "").replace("¥", "").replace("￥", "")
    try:
        return f"¥{float(raw):,.2f}"
    except ValueError:
        return text(value) or "未填写"


def room_number(value):
    value = text(value)
    match = re.search(r"(\d+)", value)
    return match.group(1) if match else (value or "未填写")


def unit_price(value, unit=""):
    label = money(value)
    unit = text(unit).replace("元", "").strip()
    if unit and not unit.startswith("/"):
        unit = f"/{unit}"
    return f"{label}{unit}" if unit else label


def contract_detail(row, indexes):
    def field(name):
        return text(row[indexes[name]]) or "未填写"
    return {
        "customer": field("租客姓名"),
        "building": field("小区/公寓"),
        "room": room_number(row[indexes["门牌号"]]),
        "leaseTerm": field("租期时长"),
        "amount": unit_price(row[indexes["租金单价"]], row[indexes["租金单位"]]),
        "signer": field("签约人"),
    }


def today_contract_details(run_dir, data_date):
    details = {"new": [], "renewal": [], "actualCheckout": []}
    seen_signing = set()
    seen_checkout = set()
    required = ("合同编号", "签约来源", "签约时间", "预退/实退", "租客姓名", "小区/公寓", "门牌号", "租期时长", "租金单价", "租金单位", "签约人")
    for filename in ("在租中合同.xlsx", "将搬入合同.xlsx", "已退租合同.xlsx"):
        ws = load_workbook(run_dir / filename, read_only=True, data_only=True).active
        headers = [cell.value for cell in ws[3]]
        indexes = {name: column(headers, name) for name in required}
        for row in ws.iter_rows(min_row=4, values_only=True):
            contract_id = norm(row[indexes["合同编号"]])
            if not contract_id:
                continue
            if contract_id not in seen_signing:
                if iso_date(row[indexes["签约时间"]]) == data_date:
                    source = norm(row[indexes["签约来源"]])
                    key = "renewal" if source in {"续租", "重签"} else "new"
                    details[key].append(contract_detail(row, indexes))
                seen_signing.add(contract_id)
            if filename == "已退租合同.xlsx" and contract_id not in seen_checkout:
                if iso_date(row[indexes["预退/实退"]]) == data_date:
                    details["actualCheckout"].append(contract_detail(row, indexes))
                seen_checkout.add(contract_id)
    return details


def today_reservation_details(path, data_date):
    ws = load_workbook(path, read_only=True, data_only=True).active
    headers = [cell.value for cell in ws[1]]
    required = ("预定ID", "姓名", "小区/公寓", "地址", "租期", "租金", "预定办理人", "录入日期")
    indexes = {name: column(headers, name) for name in required}
    details = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        reservation_id = norm(row[indexes["预定ID"]])
        if not reservation_id or reservation_id in seen:
            continue
        seen.add(reservation_id)
        if iso_date(row[indexes["录入日期"]]) != data_date:
            continue
        details.append({
            "customer": text(row[indexes["姓名"]]) or "未填写",
            "building": text(row[indexes["小区/公寓"]]) or "未填写",
            "room": room_number(re.sub(r"^.*座", "", text(row[indexes["地址"]]))),
            "leaseTerm": text(row[indexes["租期"]]) or "未填写",
            "amount": f"{money(row[indexes['租金']])}/月",
            "signer": text(row[indexes["预定办理人"]]) or "未填写",
        })
    return details


def detail_lines(title, items):
    lines = [f"**{title}（{len(items)}份）**"]
    if not items:
        return lines + ["> 今日暂无记录"]
    for number, item in enumerate(items, 1):
        lines.append(
            f"> {number}. {item['customer']}｜{item['building']}｜{item['room']}｜"
            f"{item['leaseTerm']}｜{item['amount']}｜{item['signer']}"
        )
    return lines


def inspect_contract(path, expected):
    ws = load_workbook(path, read_only=True, data_only=True).active
    headers = [text(c.value) for c in ws[3]]
    status_col = headers.index("合同状态") + 1
    rows = 0
    invalid = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(text(v) for v in row):
            continue
        rows += 1
        if text(row[status_col - 1]) != expected:
            invalid += 1
    return rows, invalid == 0, f"状态均为“{expected}”" if invalid == 0 else f"发现{invalid}条状态异常"


def inspect_house(path):
    ws = load_workbook(path, read_only=True, data_only=True).active
    headers = [text(c.value) for c in ws[3]]
    id_col = headers.index("房源ID")
    estate_col = headers.index("小区/公寓")
    lock_col = headers.index("锁房备注")
    status_col = headers.index("状态")
    rows = 0
    excluded = 0
    locked = 0
    rentable = 0
    unrentable = 0
    occupied = 0
    preordered = 0
    moving_in = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not any(text(v) for v in row):
            continue
        rows += 1
        remark = norm(row[lock_col])
        state = norm(row[status_col])
        if remark:
            locked += 1
        if state == "已出租":
            occupied += 1
        elif state == "空房可租":
            rentable += 1
        elif state == "空房不可租":
            unrentable += 1
            if any(word in remark for word in ("将搬入", "待搬入")):
                moving_in += 1
            elif any(word in remark for word in ("已预订", "已预定", "预订", "预定")):
                preordered += 1
        if text(row[id_col]) == "117492563" or "物业租赁中心（路线指引）" in text(row[estate_col]):
            excluded += 1
    ok = rows > 0 and excluded == 0
    vacancy_ok = rentable + unrentable > 0
    ok = ok and vacancy_ok
    detail = f"已删除指定排除项及空白尾行；已出租{occupied}间；空房可租{rentable}间、不可租{unrentable}间；其中已预订{preordered}间、将搬入{moving_in}间"
    return rows, ok, detail if ok else f"发现{excluded}条应排除数据或空房状态缺失", locked, rentable, unrentable, occupied, preordered, moving_in


def inspect_reservation(path):
    ws = load_workbook(path, read_only=True, data_only=True).active
    headers = [text(c.value) for c in ws[1]]
    status_col = headers.index("状态")
    rows = 0
    invalid = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(text(v) for v in row):
            continue
        rows += 1
        if text(row[status_col]) != "已付定":
            invalid += 1
    return rows, invalid == 0, "仅包含“已付定”" if invalid == 0 else f"发现{invalid}条非已付定数据"


def dashboard_brief(path):
    html = path.read_text(encoding="utf-8")
    match = re.search(r"const DATA\s*=\s*(\{.*?\});", html, re.S)
    if not match:
        raise RuntimeError("工作台数据载荷不存在")
    data = json.loads(match.group(1))
    rows = data.get("contractStats", {}).get("recentPerformance", [])
    today = next((row for row in rows if row.get("date") == data.get("dataDate")), None)
    if today is None:
        raise RuntimeError("工作台缺少当天运营简报")
    brief = {
        "date": data["dataDate"],
        "generatedAt": data.get("generatedDate", ""),
        "newCount": int(today.get("newCount", 0) or 0),
        "renewalCount": int(today.get("renewalCount", 0) or 0),
        "actualCheckoutCount": int(today.get("actualCheckoutCount", 0) or 0),
        "lockCount": int(next((row.get("lockCount", 0) for row in data.get("projectData", []) if row.get("name") == "全部房源汇总"), 0) or 0),
        "rentableVacancyCount": int(next((row.get("rentableVacancyCount", 0) for row in data.get("projectData", []) if row.get("name") == "全部房源汇总"), 0) or 0),
        "unrentableVacancyCount": int(next((row.get("unrentableVacancyCount", 0) for row in data.get("projectData", []) if row.get("name") == "全部房源汇总"), 0) or 0),
        "vacancyCount": int(next((row.get("vacancyCount", 0) for row in data.get("projectData", []) if row.get("name") == "全部房源汇总"), 0) or 0),
        "occupiedCount": int(next((row.get("occupiedCount", 0) for row in data.get("projectData", []) if row.get("name") == "全部房源汇总"), 0) or 0),
        "preorderCount": int(next((row.get("preorderCount", 0) for row in data.get("projectData", []) if row.get("name") == "全部房源汇总"), 0) or 0),
        "moveInCount": int(next((row.get("moveInCount", 0) for row in data.get("projectData", []) if row.get("name") == "全部房源汇总"), 0) or 0),
        "comprehensiveCount": int(next((row.get("comprehensiveCount", 0) for row in data.get("projectData", []) if row.get("name") == "全部房源汇总"), 0) or 0),
    }
    return brief


def main():
    run_dir = Path(sys.argv[1]).resolve()
    queue_dir = Path(sys.argv[2]).resolve()
    dashboard_path = Path(sys.argv[3]).resolve() if len(sys.argv) > 3 else Path("/opt/yuxiaor-automation/site/index.html")
    dashboard_url = sys.argv[4] if len(sys.argv) > 4 else DASHBOARD_URL
    queue_dir.mkdir(parents=True, exist_ok=True)

    checks = []
    source_lock_count = None
    source_rentable_count = None
    source_unrentable_count = None
    source_occupied_count = None
    source_preorder_count = None
    source_move_in_count = None
    for filename in FILES:
        path = run_dir / filename
        if not path.exists() or path.stat().st_size == 0:
            checks.append({"file": filename, "rows": 0, "ok": False, "detail": "文件缺失或为空"})
            continue
        try:
            if filename == "房源详情.xlsx":
                rows, ok, detail, source_lock_count, source_rentable_count, source_unrentable_count, source_occupied_count, source_preorder_count, source_move_in_count = inspect_house(path)
            elif filename == "预定合同.xlsx":
                rows, ok, detail = inspect_reservation(path)
            else:
                rows, ok, detail = inspect_contract(path, EXPECTED_STATUS[filename])
            item = {"file": filename, "rows": rows, "ok": ok, "detail": detail, "bytes": path.stat().st_size}
            if filename == "房源详情.xlsx":
                item["lockCount"] = source_lock_count
                item["rentableVacancyCount"] = source_rentable_count
                item["unrentableVacancyCount"] = source_unrentable_count
                item["occupiedCount"] = source_occupied_count
                item["preorderCount"] = source_preorder_count
                item["moveInCount"] = source_move_in_count
            checks.append(item)
        except Exception as exc:
            checks.append({"file": filename, "rows": 0, "ok": False, "detail": f"无法读取：{exc}"})

    valid = all(item["ok"] for item in checks)
    completed = datetime.now().astimezone()
    validation = {
        "valid": valid,
        "runDirectory": str(run_dir),
        "completedAt": completed.isoformat(),
        "checks": checks,
    }
    (run_dir / "validation.json").write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    if not valid:
        raise SystemExit("Export validation failed: " + json.dumps(validation, ensure_ascii=False))

    try:
        brief = dashboard_brief(dashboard_path)
        details = today_contract_details(run_dir, brief["date"])
        reservation_details = today_reservation_details(run_dir / "预定合同.xlsx", brief["date"])
    except Exception as exc:
        raise SystemExit(f"Dashboard validation failed: {exc}")

    lock_monitor_ok = source_lock_count is not None and source_lock_count == brief["lockCount"]
    validation["lockMonitoring"] = {
        "field": "锁房备注",
        "sourceCount": source_lock_count,
        "dashboardCount": brief["lockCount"],
        "ok": lock_monitor_ok,
    }
    vacancy_monitor_ok = (
        source_rentable_count == brief["rentableVacancyCount"]
        and source_unrentable_count == brief["unrentableVacancyCount"]
        and source_rentable_count + source_unrentable_count == brief["vacancyCount"]
    )
    validation["vacancyMonitoring"] = {
        "field": "状态",
        "sourceRentableCount": source_rentable_count,
        "sourceUnrentableCount": source_unrentable_count,
        "dashboardRentableCount": brief["rentableVacancyCount"],
        "dashboardUnrentableCount": brief["unrentableVacancyCount"],
        "dashboardVacancyCount": brief["vacancyCount"],
        "ok": vacancy_monitor_ok,
    }
    comprehensive_monitor_ok = (
        source_occupied_count == brief["occupiedCount"]
        and source_preorder_count == brief["preorderCount"]
        and source_move_in_count == brief["moveInCount"]
        and brief["comprehensiveCount"] == brief["occupiedCount"] + brief["preorderCount"] + brief["moveInCount"]
    )
    validation["comprehensiveMonitoring"] = {
        "definition": "已出租+空房不可租中的已预订和将搬入",
        "occupiedCount": source_occupied_count,
        "preorderCount": source_preorder_count,
        "moveInCount": source_move_in_count,
        "dashboardComprehensiveCount": brief["comprehensiveCount"],
        "ok": comprehensive_monitor_ok,
    }
    validation["valid"] = validation["valid"] and lock_monitor_ok and vacancy_monitor_ok and comprehensive_monitor_ok
    (run_dir / "validation.json").write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")
    if not lock_monitor_ok:
        raise SystemExit(f"Lock monitoring validation failed: {validation['lockMonitoring']}")
    if not vacancy_monitor_ok:
        raise SystemExit(f"Vacancy monitoring validation failed: {validation['vacancyMonitoring']}")
    if not comprehensive_monitor_ok:
        raise SystemExit(f"Comprehensive monitoring validation failed: {validation['comprehensiveMonitoring']}")

    detail_counts = {
        "newCount": len(details["new"]),
        "renewalCount": len(details["renewal"]),
        "actualCheckoutCount": len(details["actualCheckout"]),
    }
    mismatches = [key for key, value in detail_counts.items() if value != brief[key]]
    if mismatches:
        raise SystemExit(f"Today detail validation failed: {mismatches}; dashboard={brief}; details={detail_counts}")

    status_lines = [
        f"> 导出时间：{completed.strftime('%Y-%m-%d %H:%M:%S')}",
        "> 校验结果：✅ 5份文件全部通过",
        f"> 锁房监测：✅ 锁房备注字段正常，工作台识别 **{brief['lockCount']}间**，与房源表一致",
        f"> 空房监测：✅ 共 **{brief['vacancyCount']}间**（可租 **{brief['rentableVacancyCount']}间**、不可租 **{brief['unrentableVacancyCount']}间**），与房源表一致",
        f"> 综合在租：✅ **{brief['comprehensiveCount']}间**（已出租 **{brief['occupiedCount']}间**、不可租中的已预订 **{brief['preorderCount']}间**、将搬入 **{brief['moveInCount']}间**）",
        "",
    ]
    for item in checks:
        label = item["file"].removesuffix(".xlsx")
        status_lines.append(f"> {label}：**{item['rows']}条**（{item['detail']}）")

    today_lines = [
        f"**今日情况（{brief['date']}）**",
        f"> 新签 **{brief['newCount']}份** ｜ 续租 **{brief['renewalCount']}份** ｜ 实际退租 **{brief['actualCheckoutCount']}份** ｜ 预定 **{len(reservation_details)}份**",
        "",
    ]
    today_lines.extend(detail_lines("新签合同明细", details["new"]))
    today_lines.append("")
    today_lines.extend(detail_lines("续租合同明细", details["renewal"]))
    today_lines.append("")
    today_lines.extend(detail_lines("实际退租合同明细", details["actualCheckout"]))
    today_lines.append("")
    today_lines.extend(detail_lines("预定合同明细", reservation_details))
    today_lines.extend(["", f"[查看最新在线工作台]({dashboard_url})"])

    job = {
        "id": f"{completed.strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:8]}",
        "chatid": "wrki7WEAAAYzG-hYJ4delzv_Y7Us71ow",
        "statusSummary": "\n".join(status_lines),
        "todaySummary": "\n".join(today_lines),
        "summary": "\n".join(today_lines),
        "dashboard": {"url": dashboard_url, **brief},
        "todayDetails": details,
        "reservationDetails": reservation_details,
        "files": [str(run_dir / filename) for filename in FILES],
        "validation": validation,
    }
    temp = queue_dir / f".{job['id']}.tmp"
    target = queue_dir / f"{job['id']}.json"
    temp.write_text(json.dumps(job, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(temp, target)
    print(json.dumps(validation, ensure_ascii=False))


if __name__ == "__main__":
    main()


